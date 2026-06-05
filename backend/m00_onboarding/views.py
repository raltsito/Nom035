import math
import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from accounts.permissions import IsTenantAdmin

logger = logging.getLogger(__name__)
from .models import Trabajador, CicloNOM
from .serializers import (
    TrabajadorSerializer, TrabajadorCreateSerializer, TrabajadorUpdateSerializer,
    CicloNOMSerializer, CicloNOMCreateSerializer,
)
from .importador import importar_excel

_PERS_LARGO  = {'sindicalizado': 'Sindicalizado / Directo', 'confianza': 'Confianza / Indirecto',
                'salary': 'Salary / Administrativo', 'otro': 'Otro'}
_PERS_CORTO  = {'sindicalizado': 'Directo', 'confianza': 'Indirecto',
                'salary': 'Salary', 'otro': 'Otro'}

def _pers_labels(tenant):
    if tenant and getattr(tenant, 'etiquetas_cortas', False):
        return _PERS_CORTO
    return _PERS_LARGO


def _muestra(N):
    if N < 1:
        return 0
    return math.ceil((0.9604 * N) / (0.0025 * (N - 1) + 0.9604))


def _estrato(n, N, N_e):
    if not N or not N_e:
        return 0
    return math.ceil((N_e / N) * n)


def _agrupar(workers, campo, labels=None):
    counts = {}
    for w in workers:
        val = w.get(campo) or ''
        key = val if val else 'No especificado'
        counts[key] = counts.get(key, 0) + 1
    return [
        {
            'clave':  k,
            'label':  labels.get(k, k) if labels else k,
            'total':  v,
        }
        for k, v in sorted(counts.items(), key=lambda x: -x[1])
    ]


def _agrupar_rangos(workers, campo, rangos):
    counts = {r[0]: 0 for r in rangos}
    counts['No especificado'] = 0
    for w in workers:
        val = w.get(campo)
        if val is None:
            counts['No especificado'] += 1
            continue
        matched = False
        for label, lo, hi in rangos:
            if lo <= val < hi:
                counts[label] += 1
                matched = True
                break
        if not matched:
            counts['No especificado'] += 1
    return [
        {'clave': k, 'label': k, 'total': v}
        for k, v in counts.items()
        if v > 0
    ]


def _con_muestra(filas, n, N):
    return [
        {
            **f,
            'muestra': _estrato(n, N, f['total']),
            'pct': round(f['total'] / N * 100, 1) if N else 0,
        }
        for f in filas
    ]


class TrabajadorViewSet(viewsets.ModelViewSet):
    permission_classes = (IsTenantAdmin,)
    serializer_class   = TrabajadorSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_super_admin:
            qs = Trabajador.objects.all()
            tenant_id = self.request.query_params.get('tenant_id')
            if tenant_id:
                qs = qs.filter(tenant_id=tenant_id)
        else:
            qs = Trabajador.objects.filter(tenant=user.tenant)
        return qs.order_by('apellido_paterno', 'apellido_materno', 'nombre')

    def get_serializer_class(self):
        if self.action == 'create':
            return TrabajadorCreateSerializer
        if self.action in ('update', 'partial_update'):
            return TrabajadorUpdateSerializer
        return TrabajadorSerializer

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()

        q = request.query_params.get('q', '').strip()
        if q:
            qs = (
                qs.filter(nombre__icontains=q) |
                qs.filter(apellido_paterno__icontains=q) |
                qs.filter(apellido_materno__icontains=q) |
                qs.filter(num_empleado__icontains=q) |
                qs.filter(email__icontains=q) |
                qs.filter(area__icontains=q) |
                qs.filter(puesto__icontains=q)
            ).distinct()

        solo_activos = request.query_params.get('activos')
        if solo_activos == '1':
            qs = qs.filter(activo=True)

        serializer = TrabajadorSerializer(qs, many=True)
        return Response({
            'data': serializer.data,
            'meta': {'count': qs.count()},
            'errors': None,
        })

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return Response({'data': TrabajadorSerializer(instance).data, 'meta': {}, 'errors': None})

    def create(self, request, *args, **kwargs):
        serializer = TrabajadorCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            instance = serializer.save()
            return Response(
                {'data': TrabajadorSerializer(instance).data, 'meta': {}, 'errors': None},
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {'data': None, 'meta': {}, 'errors': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def update(self, request, *args, **kwargs):
        partial  = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = TrabajadorUpdateSerializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'data': TrabajadorSerializer(instance).data, 'meta': {}, 'errors': None})
        return Response(
            {'data': None, 'meta': {}, 'errors': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'data': None, 'meta': {}, 'errors': None}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        instance = self.get_object()
        instance.activo = not instance.activo
        instance.save(update_fields=['activo'])
        return Response({'data': TrabajadorSerializer(instance).data, 'meta': {}, 'errors': None})

    def resumen_muestra(self, request):
        qs = self.get_queryset().filter(activo=True)
        N  = qs.count()
        n  = _muestra(N)

        workers = list(qs.values(
            'area', 'sexo', 'tipo_contratacion',
            'tipo_personal', 'experiencia_anios', 'edad',
        ))

        SEXO = {'M': 'Masculino', 'F': 'Femenino', '': 'No especificado'}
        CONT = {'planta': 'Planta', 'eventual': 'Eventual', 'subcontratado': 'Subcontratado'}
        PERS = _pers_labels(request.user.tenant)
        EDAD_R = [
            ('18-29 años', 18, 30),
            ('30-39 años', 30, 40),
            ('40-49 años', 40, 50),
            ('50+ años',   50, 9999),
        ]
        EXP_R = [
            ('Menos de 1 año',  0,  1),
            ('1-5 años',        1,  6),
            ('6-10 años',       6, 11),
            ('Más de 10 años', 11, 9999),
        ]

        estratos = {
            'area':              _con_muestra(_agrupar(workers, 'area'),                           n, N),
            'sexo':              _con_muestra(_agrupar(workers, 'sexo', SEXO),                     n, N),
            'tipo_contratacion': _con_muestra(_agrupar(workers, 'tipo_contratacion', CONT),        n, N),
            'tipo_puesto':       _con_muestra(_agrupar(workers, 'tipo_personal', PERS),            n, N),
            'edad':              _con_muestra(_agrupar_rangos(workers, 'edad', EDAD_R),            n, N),
            'antiguedad':        _con_muestra(_agrupar_rangos(workers, 'experiencia_anios', EXP_R), n, N),
        }

        return Response({
            'data': {
                'N': N,
                'n': n,
                'pct': round(n / N * 100, 1) if N else 0,
                'estratos': estratos,
            },
            'meta':   {},
            'errors': None,
        })

    @action(
        detail=False,
        methods=['post'],
        url_path='importar-excel',
        parser_classes=[MultiPartParser],
    )
    def importar_excel(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'data': None, 'meta': {}, 'errors': {'archivo': ['Se requiere un archivo Excel.']}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext = archivo.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls'):
            return Response(
                {'data': None, 'meta': {}, 'errors': {'archivo': ['Solo se aceptan archivos .xlsx o .xls.']}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Super admin puede importar a cualquier tenant vía ?tenant_id=
        if request.user.is_super_admin:
            from django.apps import apps
            tenant_id = request.query_params.get('tenant_id') or request.data.get('tenant_id')
            if not tenant_id:
                return Response(
                    {'data': None, 'meta': {}, 'errors': {'tenant_id': ['Especifica el tenant.']}},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            Tenant = apps.get_model('accounts', 'Tenant')
            try:
                tenant = Tenant.objects.get(pk=tenant_id)
            except Tenant.DoesNotExist:
                return Response(
                    {'data': None, 'meta': {}, 'errors': {'tenant_id': ['Tenant no encontrado.']}},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            tenant = request.user.tenant

        try:
            resultado = importar_excel(archivo, tenant)
        except Exception as exc:
            return Response(
                {'data': None, 'meta': {}, 'errors': {'archivo': [f'Error al procesar el archivo: {exc}']}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'data': resultado,
            'meta': {'tenant': str(tenant)},
            'errors': None,
        }, status=status.HTTP_200_OK)


class CicloNOMViewSet(viewsets.ModelViewSet):
    permission_classes = (IsTenantAdmin,)
    serializer_class   = CicloNOMSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_super_admin:
            return CicloNOM.objects.all()
        return CicloNOM.objects.filter(tenant=user.tenant)

    def get_serializer_class(self):
        if self.action == 'create':
            return CicloNOMCreateSerializer
        return CicloNOMSerializer

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        serializer = CicloNOMSerializer(qs, many=True)
        return Response({
            'data': serializer.data,
            'meta': {'count': qs.count()},
            'errors': None,
        })

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return Response({'data': CicloNOMSerializer(instance).data, 'meta': {}, 'errors': None})

    def create(self, request, *args, **kwargs):
        serializer = CicloNOMCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            instance = serializer.save()
            return Response(
                {'data': CicloNOMSerializer(instance).data, 'meta': {}, 'errors': None},
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {'data': None, 'meta': {}, 'errors': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def update(self, request, *args, **kwargs):
        partial  = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = CicloNOMSerializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({'data': CicloNOMSerializer(instance).data, 'meta': {}, 'errors': None})
        return Response(
            {'data': None, 'meta': {}, 'errors': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'data': None, 'meta': {}, 'errors': None}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avanzar-estado')
    def avanzar_estado(self, request, pk=None):
        instance = self.get_object()
        flujo = ['iniciado', 'en_progreso', 'completado', 'cerrado']
        idx   = flujo.index(instance.estado)
        if idx < len(flujo) - 1:
            instance.estado = flujo[idx + 1]
            instance.save(update_fields=['estado'])
        return Response({'data': CicloNOMSerializer(instance).data, 'meta': {}, 'errors': None})


# ── Vista standalone para resumen de muestra ─────────────────────────────────
from rest_framework.decorators import api_view, permission_classes as drf_permission_classes

@api_view(['GET'])
@drf_permission_classes([IsTenantAdmin])
def resumen_muestra_view(request):
    user = request.user
    tenant = user.tenant
    if user.is_super_admin:
        qs = Trabajador.objects.all()
        tenant_id = request.query_params.get('tenant_id')
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)
            from tenants.models import Tenant as TenantModel
            tenant = TenantModel.objects.filter(id=tenant_id).first()
        else:
            tenant = None
    else:
        qs = Trabajador.objects.filter(tenant=user.tenant)

    qs = qs.filter(activo=True)
    N  = qs.count()
    n  = _muestra(N)

    workers = list(qs.values(
        'area', 'sexo', 'tipo_contratacion',
        'tipo_personal', 'experiencia_anios', 'edad',
    ))

    SEXO = {'M': 'Masculino', 'F': 'Femenino'}
    CONT = {'planta': 'Planta', 'eventual': 'Eventual', 'subcontratado': 'Subcontratado'}
    PERS = _pers_labels(tenant)
    EDAD_R = [
        ('18-29 años', 18, 30),
        ('30-39 años', 30, 40),
        ('40-49 años', 40, 50),
        ('50+ años',   50, 9999),
    ]
    EXP_R = [
        ('Menos de 1 año',  0,  1),
        ('1-5 años',        1,  6),
        ('6-10 años',       6, 11),
        ('Más de 10 años', 11, 9999),
    ]

    estratos = {
        'area':              _con_muestra(_agrupar(workers, 'area'),                            n, N),
        'sexo':              _con_muestra(_agrupar(workers, 'sexo', SEXO),                      n, N),
        'tipo_contratacion': _con_muestra(_agrupar(workers, 'tipo_contratacion', CONT),         n, N),
        'tipo_puesto':       _con_muestra(_agrupar(workers, 'tipo_personal', PERS),             n, N),
        'edad':              _con_muestra(_agrupar_rangos(workers, 'edad', EDAD_R),             n, N),
        'antiguedad':        _con_muestra(_agrupar_rangos(workers, 'experiencia_anios', EXP_R), n, N),
    }

    return Response({
        'data': {
            'N': N,
            'n': n,
            'pct': round(n / N * 100, 1) if N else 0,
            'estratos': estratos,
        },
        'meta':   {},
        'errors': None,
    })


import random
import io
from datetime import date as _date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill,
)
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# ── Paleta ────────────────────────────────────────────────────────────────────
_ACCENT   = '03C4CE'   # teal principal
_NAVY     = '0A2540'   # azul oscuro para títulos
_LIGHT    = 'E8FAFB'   # fondo alterno claro
_WHITE    = 'FFFFFF'
_GRAY     = 'F4F6F8'   # fondo headers de estrato
_SUBHEAD  = '64748B'   # texto gris medio

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color=_NAVY, size=10):
    return Font(bold=bold, color=color, name='Calibri', size=size)

def _border():
    s = Side(style='thin', color='DDE1E7')
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)


SEXO_LABEL   = {'M': 'Masculino', 'F': 'Femenino'}
CONT_LABEL   = {'planta': 'Planta', 'eventual': 'Eventual', 'subcontratado': 'Subcontratado'}


@api_view(['GET'])
@drf_permission_classes([IsTenantAdmin])
def sugeridos_muestra_view(request):
    user = request.user
    tenant = user.tenant
    if user.is_super_admin:
        qs = Trabajador.objects.all()
        tenant_id = request.query_params.get('tenant_id')
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)
            from tenants.models import Tenant as TenantModel
            tenant = TenantModel.objects.filter(id=tenant_id).first()
        else:
            tenant = None
    else:
        qs = Trabajador.objects.filter(tenant=user.tenant)

    qs = qs.filter(activo=True)
    N  = qs.count()
    n  = _muestra(N)

    todos = list(qs.values(
        'id', 'nombre', 'apellido_paterno', 'apellido_materno',
        'num_empleado', 'area', 'puesto',
        'sexo', 'edad', 'tipo_contratacion', 'tipo_personal', 'experiencia_anios',
    ))

    # Agrupar por área (estratificación primaria)
    grupos = {}
    for w in todos:
        key = w['area'] or 'Sin área'
        grupos.setdefault(key, []).append(w)

    seed = int(request.query_params.get('seed', 1))
    rng  = random.Random(seed)

    seleccionados = []
    for area, miembros in sorted(grupos.items()):
        n_area = _estrato(n, N, len(miembros))
        elegidos = rng.sample(miembros, min(n_area, len(miembros)))
        for w in elegidos:
            seleccionados.append({
                'id':              w['id'],
                'num_empleado':    w['num_empleado'] or '—',
                'nombre_completo': ' '.join(filter(None, [w['nombre'], w['apellido_paterno'], w['apellido_materno']])),
                'area':            w['area'] or 'Sin área',
                'puesto':          w['puesto'] or '—',
                'sexo':            SEXO_LABEL.get(w['sexo'], '—'),
                'edad':            w['edad'],
                'tipo_contratacion': CONT_LABEL.get(w['tipo_contratacion'], '—'),
                'tipo_personal':   _pers_labels(tenant).get(w['tipo_personal'], '—'),
                'experiencia_anios': w['experiencia_anios'],
            })

    seleccionados.sort(key=lambda w: (w['area'], w['nombre_completo']))

    return Response({
        'data': {
            'n': n,
            'N': N,
            'seed': seed,
            'trabajadores': seleccionados,
        },
        'meta':   {'count': len(seleccionados)},
        'errors': None,
    })


@api_view(['GET'])
@drf_permission_classes([IsTenantAdmin])
def exportar_muestra_excel(request):
    try:
        return _exportar_muestra_excel_impl(request)
    except Exception as exc:
        logger.exception('exportar_muestra_excel falló: %s', exc)
        return Response(
            {'data': None, 'meta': {}, 'errors': {'detail': 'Error al generar el archivo Excel.'}},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def _exportar_muestra_excel_impl(request):
    user = request.user
    tenant = user.tenant
    if user.is_super_admin:
        qs = Trabajador.objects.all()
        tenant_id = request.query_params.get('tenant_id')
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)
            from tenants.models import Tenant as TenantModel
            tenant = TenantModel.objects.filter(id=tenant_id).first()
        else:
            tenant = None
        tenant_nombre = 'Todas las empresas'
    else:
        qs = Trabajador.objects.filter(tenant=user.tenant)
        tenant_nombre = str(user.tenant) if user.tenant else '—'

    qs   = qs.filter(activo=True)
    N    = qs.count()
    n    = _muestra(N)
    seed = int(request.query_params.get('seed', 1))
    rng  = random.Random(seed)

    workers_raw = list(qs.values(
        'id', 'nombre', 'apellido_paterno', 'apellido_materno',
        'num_empleado', 'area', 'puesto',
        'sexo', 'edad', 'tipo_contratacion', 'tipo_personal', 'experiencia_anios',
    ))

    # ── Calcular estratos ────────────────────────────────────────────────────
    SEXO_L = {'M': 'Masculino', 'F': 'Femenino'}
    CONT_L = {'planta': 'Planta', 'eventual': 'Eventual', 'subcontratado': 'Subcontratado'}
    PERS_L = _pers_labels(tenant)
    EDAD_R = [('18-29 años', 18, 30), ('30-39 años', 30, 40), ('40-49 años', 40, 50), ('50+ años', 50, 9999)]
    EXP_R  = [('Menos de 1 año', 0, 1), ('1-5 años', 1, 6), ('6-10 años', 6, 11), ('Más de 10 años', 11, 9999)]

    estratos_cfg = [
        ('Área / Departamento',   _con_muestra(_agrupar(workers_raw, 'area'), n, N)),
        ('Sexo',                  _con_muestra(_agrupar(workers_raw, 'sexo', SEXO_L), n, N)),
        ('Tipo de contratación',  _con_muestra(_agrupar(workers_raw, 'tipo_contratacion', CONT_L), n, N)),
        ('Tipo de puesto',        _con_muestra(_agrupar(workers_raw, 'tipo_personal', PERS_L), n, N)),
        ('Edad',                  _con_muestra(_agrupar_rangos(workers_raw, 'edad', EDAD_R), n, N)),
        ('Antigüedad',            _con_muestra(_agrupar_rangos(workers_raw, 'experiencia_anios', EXP_R), n, N)),
    ]

    # ── Sugeridos ────────────────────────────────────────────────────────────
    grupos = {}
    for w in workers_raw:
        grupos.setdefault(w['area'] or 'Sin área', []).append(w)

    sugeridos = []
    for area, miembros in sorted(grupos.items()):
        elegidos = rng.sample(miembros, min(_estrato(n, N, len(miembros)), len(miembros)))
        for w in elegidos:
            sugeridos.append({
                'num_empleado':    w['num_empleado'] or '—',
                'nombre_completo': ' '.join(filter(None, [w['nombre'], w['apellido_paterno'], w['apellido_materno']])),
                'area':            w['area'] or 'Sin área',
                'puesto':          w['puesto'] or '—',
                'sexo':            SEXO_L.get(w['sexo'] or '', '—'),
                'edad':            w['edad'] if w['edad'] is not None else '—',
                'tipo_contratacion': CONT_L.get(w['tipo_contratacion'] or '', '—'),
                'tipo_personal':   PERS_L.get(w['tipo_personal'] or '', '—'),
                'experiencia_anios': w['experiencia_anios'] if w['experiencia_anios'] is not None else '—',
            })
    sugeridos.sort(key=lambda w: (w['area'], w['nombre_completo']))

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════
    # Hoja 1 — Resumen de muestra
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen de Muestra'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18

    def _h(row, col, val, **kw):
        c = ws1.cell(row=row, column=col, value=val)
        for k, v in kw.items():
            setattr(c, k, v)
        return c

    # Fila 1: título principal
    ws1.merge_cells('A1:D1')
    c = ws1.cell(row=1, column=1, value=f'Calculadora de Muestra NOM-035-STPS-2018')
    c.fill      = _fill(_NAVY)
    c.font      = Font(bold=True, color=_WHITE, name='Calibri', size=14)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 32

    # Fila 2: empresa y fecha
    ws1.merge_cells('A2:D2')
    c = ws1.cell(row=2, column=1, value=f'{tenant_nombre}   ·   Generado: {_date.today().strftime("%d/%m/%Y")}')
    c.fill      = _fill(_ACCENT)
    c.font      = Font(bold=False, color=_WHITE, name='Calibri', size=10)
    c.alignment = _center()
    ws1.row_dimensions[2].height = 22

    # Fila 3: vacía
    ws1.row_dimensions[3].height = 10

    # Filas 4-5: KPIs
    kpis = [
        ('Total trabajadores (N)', N),
        ('Muestra requerida (n)',  n),
        ('% de la plantilla',     f'{round(n/N*100,1)}%' if N else '—'),
        ('Metodología',           'Confianza 95% · Error ±5%'),
    ]
    for col, (lbl, val) in enumerate(kpis, start=1):
        lc = ws1.cell(row=4, column=col, value=lbl)
        lc.fill      = _fill(_ACCENT)
        lc.font      = Font(bold=True, color=_WHITE, name='Calibri', size=9)
        lc.alignment = _center(wrap=True)
        lc.border    = _border()
        vc = ws1.cell(row=5, column=col, value=val)
        vc.fill      = _fill(_LIGHT)
        vc.font      = Font(bold=True, color=_NAVY, name='Calibri', size=13)
        vc.alignment = _center()
        vc.border    = _border()
    ws1.row_dimensions[4].height = 22
    ws1.row_dimensions[5].height = 28

    row = 7
    for nombre_estrato, filas in estratos_cfg:
        # Encabezado de sección
        ws1.merge_cells(f'A{row}:D{row}')
        c = ws1.cell(row=row, column=1, value=f'  {nombre_estrato}')
        c.fill      = _fill(_NAVY)
        c.font      = Font(bold=True, color=_WHITE, name='Calibri', size=10)
        c.alignment = _left()
        ws1.row_dimensions[row].height = 20
        row += 1

        # Sub-cabeceras
        for col, txt in enumerate(['Grupo', 'Trabajadores', '% del total', 'Muestra sugerida'], 1):
            c = ws1.cell(row=row, column=col, value=txt)
            c.fill      = _fill(_ACCENT)
            c.font      = Font(bold=True, color=_WHITE, name='Calibri', size=9)
            c.alignment = _center()
            c.border    = _border()
        ws1.row_dimensions[row].height = 18
        row += 1

        # Datos
        for i, f in enumerate(filas):
            bg = _LIGHT if i % 2 == 0 else _WHITE
            cells = [
                (f['label'],                               _left()),
                (f['total'],                               _center()),
                (f'{f["pct"]}%',                           _center()),
                (f['muestra'],                             _center()),
            ]
            for col, (val, aln) in enumerate(cells, 1):
                c = ws1.cell(row=row, column=col, value=val)
                c.fill      = _fill(bg)
                c.font      = _font(bold=(col == 4), color=_ACCENT if col == 4 else _NAVY)
                c.alignment = aln
                c.border    = _border()
            row += 1

        row += 1  # espacio entre secciones

    # ════════════════════════════════════════════════════════
    # Hoja 2 — Lista de trabajadores
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Lista para Encuestar')
    ws2.sheet_view.showGridLines = False

    COLS2 = [
        ('#',                 5),
        ('No. Empleado',     14),
        ('Nombre completo',  32),
        ('Área',             22),
        ('Puesto',           24),
        ('Sexo',             12),
        ('Edad',              8),
        ('Contratación',     18),
        ('Tipo personal',    22),
        ('Antigüedad (años)', 16),
    ]
    for i, (_, w) in enumerate(COLS2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws2.merge_cells(f'A1:{get_column_letter(len(COLS2))}1')
    c = ws2.cell(row=1, column=1, value=f'Lista de trabajadores a encuestar — {tenant_nombre}  ·  {_date.today().strftime("%d/%m/%Y")}')
    c.fill      = _fill(_NAVY)
    c.font      = Font(bold=True, color=_WHITE, name='Calibri', size=12)
    c.alignment = _center()
    ws2.row_dimensions[1].height = 28

    c2 = ws2.cell(row=2, column=1, value=f'Muestra representativa: {n} de {N} trabajadores  ·  Selección aleatoria estratificada por área')
    ws2.merge_cells(f'A2:{get_column_letter(len(COLS2))}2')
    c2.fill      = _fill(_ACCENT)
    c2.font      = Font(bold=False, color=_WHITE, name='Calibri', size=10)
    c2.alignment = _center()
    ws2.row_dimensions[2].height = 20

    # Cabecera columnas
    for col, (lbl, _) in enumerate(COLS2, 1):
        c = ws2.cell(row=3, column=col, value=lbl)
        c.fill      = _fill(_NAVY)
        c.font      = Font(bold=True, color=_WHITE, name='Calibri', size=9)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    ws2.row_dimensions[3].height = 22

    for i, w in enumerate(sugeridos, 1):
        bg = _LIGHT if i % 2 == 0 else _WHITE
        row_data = [
            i, w['num_empleado'], w['nombre_completo'], w['area'],
            w['puesto'], w['sexo'], w['edad'],
            w['tipo_contratacion'], w['tipo_personal'], w['experiencia_anios'],
        ]
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=3 + i, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(col == 3), color=_NAVY if col != 1 else _SUBHEAD)
            c.alignment = _center() if col in (1, 6, 7) else _left()
            c.border    = _border()
        ws2.row_dimensions[3 + i].height = 17

    # ── Respuesta ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'muestra_nom035_{_date.today().isoformat()}.xlsx'
    resp  = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
