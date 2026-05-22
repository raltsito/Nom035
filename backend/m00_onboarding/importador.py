"""
Importador de Excel para padrón de trabajadores (NOM-035).

Formato estándar (columnas numeradas por Benja):
  ID Trabajador | Nombre / Folio | 1. Sexo | 2. Edad (Años) | 3. Estado Civil |
  4. Nivel de Estudios | 5. Puesto / Profesión | 6. Departamento / Área |
  7. Tipo de Puesto | 8. Tipo de Contratación | 9. Tipo de Personal |
  10. Jornada de Trabajo | 11. ¿Rota Turnos? | 12. Exp. Puesto Actual (Años) |
  13. Exp. Empresa Actual (Años) | 14. Exp. Laboral (Años)

También soporta los formatos heredados de plantas Lear (columnas con nombres
distintos, nombres en una o tres columnas, valores numéricos para sexo/jornada).
"""

import re
import unicodedata
from io import BytesIO

import openpyxl

from .models import Trabajador


# ---------------------------------------------------------------------------
# Normalización de texto
# ---------------------------------------------------------------------------

def _norm(value):
    """Lowercase sin acentos, signos de puntuación especiales ni espacios extra."""
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r'[¿?¡!]', '', text)   # eliminar signos de interrogación/exclamación
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _clean(value):
    if value is None:
        return ''
    return str(value).strip()


# ---------------------------------------------------------------------------
# Mapeo de encabezados → campo canónico
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    'num_empleado': [
        # Estándar nuevo
        'id trabajador',
        # Formatos heredados
        'codigoemp', 'codigo', 'no. empleado', 'no empleado', 'num empleado',
        'numero de empleado', 'numero empleado', '#', 'cod emp',
        'employee id', 'id empleado',
    ],
    'nombre_completo': [
        # Estándar nuevo
        'nombre / folio',
        # Formatos heredados
        'nombre del empleado',
    ],
    'nombre': [
        'nombre', 'nombres', 'nombre(s)',
    ],
    'apellido_paterno': [
        'apellido paterno', 'apellido', 'ap. paterno', 'ap paterno',
        'apellido1', 'primer apellido',
    ],
    'apellido_materno': [
        'apellido materno', 'apellido2', 'ap. materno', 'ap materno',
        'segundo apellido',
    ],
    'sexo': [
        # Estándar nuevo (con prefijo numérico)
        '1. sexo',
        # Formatos heredados
        'sexo', 'genero', 'genero/sexo',
    ],
    'edad': [
        '2. edad (anos)',
        'edad',
    ],
    'estado_civil': [
        '3. estado civil',
        'estado civil', 'estado civil.', 'estado_civil',
    ],
    'nivel_estudios': [
        '4. nivel de estudios',
        'nivel de estudios', 'nivel estudios', 'escolaridad',
        'nivel de escolaridad',
    ],
    'puesto': [
        '5. puesto / profesion',
        'puesto', 'ocupacion / profesion / puesto', 'ocupacion/profesion/puesto',
        'puesto / cargo', 'cargo', 'ocupacion', 'profesion',
        'tipo de puesto', 'tipo puesto',
    ],
    'area': [
        '6. departamento / area',
        'departamento', 'departamento/ seccion/ area',
        'departamento/seccion/area', 'area', 'seccion',
        'area / departamento', 'departamento/area', 'area/departamento',
        'direccion departamento',
    ],
    # Columna 7 "Tipo de Puesto" del estándar nuevo es redundante con puesto/personal,
    # se omite intencionalmente para no pisar col 5.
    'tipo_contratacion': [
        '8. tipo de contratacion',
        'tipo de contratacion', 'tipo contratacion',
    ],
    'tipo_personal': [
        '9. tipo de personal',
        'tipo de personal', 'tipo personal',
    ],
    'tipo_jornada': [
        '10. jornada de trabajo',
        'tipo jornada de trabajo', 'tipo de jornada de trabajo',
        'tipo de jornada', 'tipo jornada', 'turno', 'jornada',
    ],
    'rotacion_turnos': [
        '11. rota turnos',
        'realiza rotacion de turnos', 'realiza rotacion de turno',
        'rotacion de turnos', 'rotacion de turno', 'rotacion turnos',
    ],
    'tiempo_puesto_actual': [
        '12. exp. puesto actual (anos)',
        'tiempo en el puesto actual', 'tiempo en puesto actual',
        'tiempo de puesto actual', 'tiempo en puesto',
        'tiempo puesto actual', 'tiempo en el puesto',
    ],
    'experiencia_empresa_anios': [
        '13. exp. empresa actual (anos)',
    ],
    'experiencia_anios': [
        '14. exp. laboral (anos)',
        'experiencia en anos', 'experiencia en a;os',
        'experiencia', 'experinecia en a;os',
        'experiencia laboral', 'anos de experiencia', 'antiguedad',
    ],
}

_ALIAS_INDEX = {}
for _campo, _aliases in COLUMN_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_INDEX[_norm(_alias)] = _campo


def _map_column(header):
    return _ALIAS_INDEX.get(_norm(header))


# ---------------------------------------------------------------------------
# Normalización de valores
# ---------------------------------------------------------------------------

def _norm_sexo(value):
    v = _norm(value)
    if v in ('masculino', 'hombre', 'm', 'male', 'h'):
        return 'M'
    if v in ('femenino', 'mujer', 'f', 'female'):
        return 'F'
    return ''


def _norm_estado_civil(value):
    v = _norm(value)
    if not v or v == 'none':
        return ''
    # Descartar valores que parecen edades (bug en Querétaro)
    try:
        if int(float(value)) < 100:
            return ''
    except (ValueError, TypeError):
        pass
    if 'soltero' in v or 'soltera' in v:
        return 'soltero'
    if 'casado' in v or 'casada' in v:
        return 'casado'
    if 'union' in v or 'libre' in v or 'concubinato' in v:
        return 'union_libre'
    if 'divorciado' in v or 'divorciada' in v or 'separado' in v:
        return 'divorciado'
    if 'viudo' in v or 'viuda' in v:
        return 'viudo'
    try:
        code = int(float(value))
        return {1: 'soltero', 2: 'casado', 3: 'divorciado', 4: 'viudo', 5: 'union_libre'}.get(code, 'otro')
    except (ValueError, TypeError):
        pass
    return 'otro'


def _norm_nivel_estudios(value):
    v = _norm(value)
    if not v or v == 'none':
        return ''
    if 'primaria' in v:
        return 'primaria'
    if 'secundaria' in v:
        return 'secundaria'
    if 'bachillerato' in v or 'preparatoria' in v or 'prepa' in v or 'media superior' in v:
        return 'bachillerato'
    if 'tsu' in v or 'tecnico superior' in v or 'tecnica superior' in v:
        return 'tecnico'
    if 'tecnico' in v or 'tecnica' in v or 'carrera tecnica' in v:
        return 'tecnico'
    if 'ingenieria' in v or 'ingenieria' in v:
        return 'ingenieria'
    if 'licenciatura' in v or 'profesional' in v:
        return 'licenciatura'
    if 'maestria' in v or 'doctorado' in v or 'posgrado' in v or 'postgrado' in v or 'maestr' in v:
        return 'posgrado'
    return 'otro'


def _norm_tipo_personal(value):
    v = _norm(value)
    if not v or v == 'none':
        return ''
    if 'sindicalizado' in v or 'directo' in v or 'hourly sind' in v:
        return 'sindicalizado'
    if 'indirecto' in v or 'confianza' in v or 'hourly no sind' in v:
        return 'confianza'
    if 'salary' in v or 'administrativo' in v or 'hourly admin' in v or \
       'asalariado' in v or 'no sindicalizado' in v:
        return 'salary'
    try:
        code = int(float(value))
        return {1: 'sindicalizado', 2: 'confianza', 3: 'sindicalizado',
                5: 'salary', 8: 'confianza'}.get(code, 'otro')
    except (ValueError, TypeError):
        pass
    return 'otro'


def _norm_tipo_jornada(value):
    v = _norm(value)
    if not v or v == 'none':
        return ''
    # Texto descriptivo
    if 'diurno' in v or 'diurna' in v or 'primer turno' in v:
        return 'diurno'
    if 'nocturno' in v or 'nocturna' in v or 'tercer turno' in v:
        return 'nocturno'
    if 'mixto' in v or 'mixta' in v or 'segundo turno' in v:
        return 'mixto'
    # Códigos numéricos: 1=Diurno, 2=Nocturno, 3=Mixto (convención Lear)
    try:
        code = int(float(value))
        return {1: 'diurno', 2: 'nocturno', 3: 'mixto', 4: 'nocturno'}.get(code, '')
    except (ValueError, TypeError):
        pass
    return ''


def _norm_rotacion(value):
    v = _norm(value)
    if not v or v == 'none':
        return None
    if v in ('si', 'sí', 'yes', 's', '1', 'true'):
        return True
    if v in ('no', 'n', '0', 'false'):
        return False
    return None


def _norm_edad(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        pass
    v = str(value)
    match = re.search(r'(\d+)\s*a', v, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def _norm_experiencia(value):
    if value is None:
        return None
    v = str(value).strip()
    try:
        return round(float(v), 1)
    except (ValueError, TypeError):
        pass
    match = re.search(r'(\d+)\s*a', v, re.IGNORECASE)
    meses = re.search(r'(\d+)\s*m', v, re.IGNORECASE)
    if match:
        years = int(match.group(1))
        months = int(meses.group(1)) if meses else 0
        return round(years + months / 12, 1)
    return None


def _norm_contratacion(value):
    v = _norm(value)
    if 'indeterminado' in v or 'individual' in v or 'indefinido' in v or \
       'colectivo' in v or 'planta' in v:
        return 'planta'
    if 'temporal' in v or 'determinado' in v or 'eventual' in v:
        return 'eventual'
    if 'subcontrat' in v or 'outsourc' in v:
        return 'subcontratado'
    return 'planta'


# ---------------------------------------------------------------------------
# Detección de fila de encabezados (compatibilidad con formatos heredados)
# ---------------------------------------------------------------------------

def _detect_header_row(ws, max_scan=6):
    best_row, best_score = 1, -1
    for row_idx in range(1, max_scan + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        score = sum(1 for cell in row if cell and _map_column(str(cell)) is not None)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


# ---------------------------------------------------------------------------
# Parseo de nombre completo
# ---------------------------------------------------------------------------

def _split_nombre_completo(full_name):
    """Divide 'AP_PAT AP_MAT NOMBRE(S)' o 'NOMBRE AP_PAT AP_MAT' en tres partes."""
    parts = _clean(full_name).split()
    if not parts:
        return '', '', ''
    if len(parts) == 1:
        return parts[0].title(), '', ''
    if len(parts) == 2:
        return parts[1].title(), parts[0].title(), ''
    # 3+ palabras: asumir AP_PAT AP_MAT NOMBRE(S)
    ap_pat = parts[0].title()
    ap_mat = parts[1].title()
    nombre = ' '.join(p.title() for p in parts[2:])
    return nombre, ap_pat, ap_mat


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def importar_excel(file_obj, tenant):
    """
    Parsea un Excel con el formato estándar (o heredado) y crea/actualiza
    trabajadores del tenant.

    Retorna:
        {
            'importados': int,
            'actualizados': int,
            'omitidos': int,
            'errores': [{'fila': N, 'motivo': str}, ...],
        }
    """
    content = file_obj.read() if hasattr(file_obj, 'read') else file_obj
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx = _detect_header_row(ws)

    headers_raw = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    ))[0]
    headers = [str(h).strip() if h is not None else '' for h in headers_raw]

    col_map = {}
    for idx, h in enumerate(headers):
        campo = _map_column(h)
        if campo and campo not in col_map:
            col_map[campo] = idx

    data_rows = list(ws.iter_rows(min_row=header_row_idx + 1, values_only=True))
    wb.close()

    importados = actualizados = omitidos = 0
    errores = []

    for row_num, row in enumerate(data_rows, start=header_row_idx + 1):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        def get(campo):
            idx = col_map.get(campo)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return v if v != '' else None

        try:
            # ── Nombre ──────────────────────────────────────────────────────
            if 'nombre_completo' in col_map:
                nombre, ap_pat, ap_mat = _split_nombre_completo(get('nombre_completo') or '')
            elif 'apellido_paterno' in col_map:
                nombre = _clean(get('nombre') or '').title()
                ap_pat = _clean(get('apellido_paterno') or '').title()
                ap_mat = _clean(get('apellido_materno') or '').title()
            elif 'nombre' in col_map:
                raw = get('nombre')
                try:
                    int(float(raw))
                    nombre, ap_pat, ap_mat = '', '', ''
                except (ValueError, TypeError):
                    nombre, ap_pat, ap_mat = _split_nombre_completo(raw or '')
            else:
                nombre, ap_pat, ap_mat = '', '', ''

            if not nombre and not ap_pat:
                omitidos += 1
                continue

            # ── No. empleado ─────────────────────────────────────────────────
            raw_codigo = get('num_empleado')
            if raw_codigo is None and 'nombre' in col_map:
                raw_n = get('nombre')
                try:
                    int(float(raw_n))
                    raw_codigo = raw_n
                except (ValueError, TypeError):
                    pass
            num_empleado = str(int(float(raw_codigo))) if raw_codigo is not None else ''

            # ── Campos Guía V ────────────────────────────────────────────────
            defaults = dict(
                nombre=nombre or ap_pat,
                apellido_paterno=ap_pat,
                apellido_materno=ap_mat,
                puesto=_clean(get('puesto') or '').title(),
                area=_clean(get('area') or '').title(),
                tipo_contratacion=_norm_contratacion(get('tipo_contratacion') or ''),
                sexo=_norm_sexo(get('sexo')),
                edad=_norm_edad(get('edad')),
                estado_civil=_norm_estado_civil(get('estado_civil')),
                nivel_estudios=_norm_nivel_estudios(get('nivel_estudios')),
                tipo_personal=_norm_tipo_personal(get('tipo_personal')),
                tipo_jornada=_norm_tipo_jornada(get('tipo_jornada')),
                rotacion_turnos=_norm_rotacion(get('rotacion_turnos')),
                tiempo_puesto_actual=_norm_experiencia(get('tiempo_puesto_actual')),
                experiencia_empresa_anios=_norm_experiencia(get('experiencia_empresa_anios')),
                experiencia_anios=_norm_experiencia(get('experiencia_anios')),
            )

            if num_empleado:
                obj, created = Trabajador.objects.update_or_create(
                    tenant=tenant,
                    num_empleado=num_empleado,
                    defaults=defaults,
                )
            else:
                qs = Trabajador.objects.filter(
                    tenant=tenant,
                    nombre=defaults['nombre'],
                    apellido_paterno=defaults['apellido_paterno'],
                )
                if qs.exists():
                    obj = qs.first()
                    for k, v in defaults.items():
                        setattr(obj, k, v)
                    obj.save()
                    created = False
                else:
                    obj = Trabajador.objects.create(tenant=tenant, **defaults)
                    created = True

            if created:
                importados += 1
            else:
                actualizados += 1

        except Exception as exc:
            errores.append({'fila': row_num, 'motivo': str(exc)})

    return {
        'importados': importados,
        'actualizados': actualizados,
        'omitidos': omitidos,
        'errores': errores,
    }
