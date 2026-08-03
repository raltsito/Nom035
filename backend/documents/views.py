import io
import logging
from collections import defaultdict
from datetime import date
from pathlib import Path

from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import status as drf_status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.permissions import IsSuperAdmin, IsTenantAdmin
from core import paleta_riesgo as paleta
from core import xlsx_styles as xs
from core.confidencialidad import ETIQUETA_RESERVADO, grupo_reservado, umbral_confidencialidad
from m00_onboarding.models import CicloNOM, Trabajador
from m00_onboarding.views import _muestra
from m05_questionnaires.models import Aplicacion, Pregunta, RespuestaPregunta
from m06_results import matriz as mat
from m06_results.matriz import (
    COLUMNAS_CATEGORIAS,
    COLUMNAS_DOMINIOS,
    COLUMNAS_SECCIONES_GUIA_I,
    construir_matriz,
    construir_matriz_guia_i,
)
from m06_results.models import ResultadoAplicacion, ResultadoDominio
from m06_results.scoring import (
    NOTA_DIMENSIONES,
    _CORTES_FINAL,
    _CORTES_CATEGORIA,
    _CORTES_DOMINIO,
    _CATEGORIA_DOMINIOS,
    _DOMINIOS_OFICIALES,
    _categoria_por_rangos,
)
from . import contenido_normativo as norm
from . import docx_postprocess as docx_pp
from . import graficas as graf
from . import interpretaciones as interp
from .report_data import build_report_data, datos_guia_v_por_trabajador, valor_efectivo_guia_v
from .docx_builder import build_informe_diagnostico_docx, build_reporte_psicologico_docx
from .models import ReportePsicologico

logger = logging.getLogger(__name__)

# San Luis y Zapotitlán 2026 se sirven como documento fijo (editado
# manualmente por dirección) en vez del generador dinámico.
SAN_LUIS_INFORME_2026 = (
    Path(__file__).resolve().parent
    / 'static'
    / 'documents'
    / 'informes'
    / 'san_luis_informe_diagnostico_2026.pdf'
)
ZAPOTITLAN_INFORME_2026 = (
    Path(__file__).resolve().parent
    / 'static'
    / 'documents'
    / 'informes'
    / 'zapotitlan_informe_diagnostico_2026.pdf'
)
SALTILLO_INFORME_2026 = (
    Path(__file__).resolve().parent
    / 'static'
    / 'documents'
    / 'informes'
    / 'saltillo_informe_diagnostico_2026.pdf'
)

# Razón social única del corporativo: todas las plantas pertenecen a la misma
# persona moral, por lo que §2.1 del informe no depende del tenant.
RAZON_SOCIAL_CORPORATIVA = 'CONSORCIO INDUSTRIAL MEXICANO DE AUTOPARTES S. de R.L. de C.V.'

# nombre de dominio oficial → categoría (5 grupos), derivado de la misma
# jerarquía que usa el motor de calificación.
_CATEGORIA_DE_DOMINIO = {dom: cat for cat, doms in _CATEGORIA_DOMINIOS for dom in doms}


_MESES_ES = [
    '', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
]


def _fecha_es(d):
    """Formatea una fecha en español sin depender del locale del sistema."""
    return f'{d.day} de {_MESES_ES[d.month]} de {d.year}'


def _wrap(data, meta=None, errors=None, status_code=drf_status.HTTP_200_OK):
    return Response({'data': data, 'meta': meta or {}, 'errors': errors}, status=status_code)


def _tenant_para_ciclo(request, ciclo_id):
    """Resuelve el tenant a consultar: para super admins, el del ciclo elegido
    (pueden operar sobre cualquier tenant); para tenant_admin, siempre el suyo."""
    if request.user.is_super_admin and ciclo_id:
        ciclo = CicloNOM.objects.filter(id=ciclo_id).select_related('tenant').first()
        return ciclo.tenant if ciclo else None
    return request.user.tenant

try:
    from weasyprint import HTML as WeasyHTML
    WEASYPRINT_OK = True
except Exception:
    WEASYPRINT_OK = False

# -------------------------------------------------------------------------
# Categorías de riesgo Guía III (5 niveles oficiales)
DIST_KEYS = ('nulo', 'bajo', 'medio', 'alto', 'muy_alto')

CAT_LABELS = {
    'nulo':     'Nulo / Despreciable',
    'bajo':     'Bajo',
    'medio':    'Medio',
    'alto':     'Alto',
    'muy_alto': 'Muy alto',
}

# Tabla 7, Guía de Referencia III, NOM-035-STPS-2018 — "Criterios para la
# toma de acciones". Transcripción fiel del texto oficial (DOF 23-oct-2018),
# verificada contra el PDF de la STPS el 2026-07-14. Las recomendaciones
# profesionales adicionales van SEPARADAS (sección de recomendaciones
# complementarias), nunca mezcladas con estos textos normativos.
ACCIONES_TABLA7 = [
    ('nulo',
     'El riesgo resulta despreciable por lo que no se requiere medidas adicionales.'),
    ('bajo',
     'Es necesario una mayor difusión de la política de prevención de riesgos '
     'psicosociales y programas para: la prevención de los factores de riesgo '
     'psicosocial, la promoción de un entorno organizacional favorable y la prevención '
     'de la violencia laboral.'),
    ('medio',
     'Se requiere revisar la política de prevención de riesgos psicosociales y programas '
     'para la prevención de los factores de riesgo psicosocial, la promoción de un '
     'entorno organizacional favorable y la prevención de la violencia laboral, así como '
     'reforzar su aplicación y difusión, mediante un Programa de intervención.'),
    ('alto',
     'Se requiere realizar un análisis de cada categoría y dominio, de manera que se '
     'puedan determinar las acciones de intervención apropiadas a través de un Programa '
     'de intervención, que podrá incluir una evaluación específica y deberá incluir una '
     'campaña de sensibilización, revisar la política de prevención de riesgos '
     'psicosociales y programas para la prevención de los factores de riesgo psicosocial, '
     'la promoción de un entorno organizacional favorable y la prevención de la violencia '
     'laboral, así como reforzar su aplicación y difusión.'),
    ('muy_alto',
     'Se requiere realizar el análisis de cada categoría y dominio para establecer las '
     'acciones de intervención apropiadas, mediante un Programa de intervención que '
     'deberá incluir evaluaciones específicas, y contemplar campañas de sensibilización, '
     'revisar la política de prevención de riesgos psicosociales y programas para la '
     'prevención de los factores de riesgo psicosocial, la promoción de un entorno '
     'organizacional favorable y la prevención de la violencia laboral, así como reforzar '
     'su aplicación y difusión.'),
]

# Alias retro-compatible (nombre anterior, incorrecto: la NOM no tiene
# "Cuadro 2"; los criterios de acción son la Tabla 7 de la Guía III).
ACCIONES_CUADRO2 = ACCIONES_TABLA7

NOTA_PROMEDIO = (
    'El puntaje promedio es un estadístico descriptivo y no constituye un nivel de '
    'riesgo oficial del centro de trabajo conforme a la NOM-035-STPS-2018. Los puntos '
    'de corte oficiales se aplican a cada cuestionario individual; la conclusión '
    'organizacional se basa en la distribución de personas por nivel.'
)

# Interpretación narrativa del nivel de riesgo global de la organización
NIVEL_GLOBAL_TEXTO = {
    'nulo':
        'La organización presenta una exposición nula o despreciable a factores de riesgo '
        'psicosocial. Se recomienda mantener las prácticas actuales y reaplicar el '
        'cuestionario en el siguiente ciclo normativo.',
    'bajo':
        'Se detecta una exposición baja a factores de riesgo psicosocial. Existen áreas de '
        'mejora que conviene atender de forma preventiva para evitar que escalen en ciclos '
        'posteriores.',
    'medio':
        'Se detecta una exposición media a factores de riesgo psicosocial. Conforme al '
        'Apartado 8 de la NOM-035-STPS-2018, el patrón debe elaborar un plan de acción con '
        'medidas concretas, plazos y responsables asignados.',
    'alto':
        'Se detecta una exposición alta a factores de riesgo psicosocial. Se requieren '
        'acciones inmediatas de intervención organizacional, seguimiento mensual y '
        'evaluación individual de los trabajadores con puntajes más elevados.',
    'muy_alto':
        'Se detecta una exposición muy alta a factores de riesgo psicosocial. Constituye '
        'una situación crítica que requiere intervención urgente, atención médica y '
        'psicológica, y un programa de intervención con indicadores de seguimiento.',
}

# Descripciones conforme a las propias guías y a los numerales 7.1-7.4 de la
# NOM-035. Nota metodológica: para centros de 16 a 50 trabajadores la NOM
# ejemplifica la identificación de FRPS con la Guía II; este sistema aplica la
# Guía III (que además evalúa entorno organizacional), lo cual cubre y excede
# ese requisito — decisión documentada en docs/auditoria_metodologica_nom035.md.
GUIA_DESC = {
    'I':   'identificación de trabajadores sujetos a acontecimientos traumáticos severos (todos los centros de trabajo)',
    'III': 'identificación de factores de riesgo psicosocial y evaluación del entorno organizacional (obligatoria en centros con más de 50 trabajadores)',
    'V':   'datos sociodemográficos y laborales del trabajador (no genera nivel de riesgo)',
}

RECOMENDACIONES_BASE = {
    'bajo': [
        ('Mantener y fortalecer las condiciones actuales del entorno organizacional.',
         'Medidas de control preventivo'),
        ('Continuar con programas de bienestar y comunicación interna.',
         'Buenas prácticas organizacionales'),
    ],
    'medio': [
        ('Revisar la distribución de carga de trabajo y ajustar en áreas de mayor presión.',
         'Factores propios de la actividad'),
        ('Implementar o reforzar canales de retroalimentación entre líderes y equipos.',
         'Liderazgo y relaciones en el trabajo'),
        ('Promover el uso de mecanismos de denuncia y atención a violencia laboral.',
         'Entorno organizacional'),
    ],
    'alto': [
        ('Diseñar e implementar un programa de intervención para reducir los niveles de riesgo identificados.',
         'Plan de acción correctivo'),
        ('Revisar las jornadas de trabajo y los mecanismos de control del tiempo.',
         'Organización del tiempo de trabajo'),
        ('Capacitar a mandos medios y supervisores en prevención de riesgos psicosociales.',
         'Liderazgo y relaciones en el trabajo'),
        ('Establecer un comité de seguridad y salud en el trabajo con seguimiento periódico.',
         'Estructura organizacional'),
    ],
    'muy_alto': [
        ('Activar de forma inmediata un programa de intervención con apoyo de especialistas.',
         'Intervención urgente'),
        ('Evaluar condiciones específicas de trabajo con personal de salud ocupacional.',
         'Salud ocupacional'),
        ('Implementar medidas de protección y apoyo psicológico para trabajadores en riesgo.',
         'Bienestar del trabajador'),
        ('Revisar políticas de contratación, estabilidad laboral y compensación.',
         'Entorno organizacional y seguridad laboral'),
        ('Documentar y dar seguimiento mensual a los indicadores de riesgo identificados.',
         'Seguimiento y control'),
    ],
}


def _get_recomendaciones(distribucion: dict) -> list:
    """Selecciona recomendaciones según el nivel de riesgo más alto encontrado."""
    recs = []
    for nivel in ('muy_alto', 'alto', 'medio', 'bajo'):
        if distribucion.get(nivel, 0) > 0:
            for texto, dominio in RECOMENDACIONES_BASE.get(nivel, []):
                recs.append({'nivel': nivel, 'texto': texto, 'dominio': dominio})
            break  # solo el nivel más alto
    # Siempre incluir la de bajo si hay riesgo bajo también
    if distribucion.get('bajo', 0) > 0 and recs and recs[0]['nivel'] != 'bajo':
        for texto, dominio in RECOMENDACIONES_BASE['bajo']:
            recs.append({'nivel': 'bajo', 'texto': texto, 'dominio': dominio})
    return recs


def _clave_map(resultados) -> dict:
    """Asigna el folio de identificación por trabajador: su número de
    empleado (Guía V, importado vía Excel — `Trabajador.num_empleado`). Si
    algún trabajador no tiene número de empleado capturado, se usa un folio
    genérico T001, T002… como respaldo, ordenado por nombre para que sea
    estable."""
    trabajadores = {r.aplicacion.trabajador_id: r.aplicacion.trabajador for r in resultados}
    ordenados = sorted(trabajadores, key=lambda tid: trabajadores[tid].nombre_completo)
    claves = {}
    for idx, tid in enumerate(ordenados, 1):
        num = trabajadores[tid].num_empleado
        claves[tid] = num if num else f'T{idx:03d}'
    return claves


def _formato_rangos(cortes_dict, orden):
    """Convierte los cortes oficiales (max_exclusive, nivel) de
    `m06_results.scoring` en filas de rango legibles para la tabla de
    referencia de la §9.2/§9.3 (ej. 'Nulo': '0-4', 'Muy alto': '15+')."""
    filas = []
    for nombre in orden:
        lo = 0
        rangos = {}
        for max_excl, nivel in cortes_dict[nombre]:
            rangos[nivel] = f'{lo}-{max_excl - 1}'
            lo = max_excl
        rangos['muy_alto'] = f'{lo}+'
        filas.append({'nombre': nombre, 'rangos': rangos})
    return filas


def _build_context(tenant, ciclo, resultados_qs, anonimo=False) -> dict:
    resultados = list(
        resultados_qs.select_related(
            'aplicacion__trabajador',
            'aplicacion__cuestionario',
        ).prefetch_related('dominios__dominio', 'dominios_oficiales')
    )

    # Población analítica: SOLO cuestionarios válidos (los 'sin_calificar' /
    # requiere_revision no entran a ninguna distribución normativa).
    res_iii = [r for r in resultados
               if r.aplicacion.cuestionario.clave == 'III'
               and getattr(r, 'estatus_validacion', 'valido') == 'valido']
    res_i   = [r for r in resultados
               if r.aplicacion.cuestionario.clave == 'I'
               and getattr(r, 'estatus_validacion', 'valido') == 'valido']
    excluidos_revision = sum(
        1 for r in resultados if getattr(r, 'estatus_validacion', 'valido') != 'valido')

    clave_map = _clave_map(resultados) if anonimo else {}

    def _nombre(r):
        if anonimo:
            return clave_map.get(r.aplicacion.trabajador_id, 'T—')
        return r.aplicacion.trabajador.nombre_completo

    total_apl  = ciclo.aplicaciones.count() if hasattr(ciclo, 'aplicaciones') else 0
    total_res  = len(resultados)
    total_comp = total_res  # todos los del queryset ya tienen resultado calculado

    # ------------------------------------------------------------------
    # Guía III — distribución global (5 niveles) y nivel de riesgo global
    # ------------------------------------------------------------------
    dist_count = {k: 0 for k in DIST_KEYS}
    for r in res_iii:
        if r.categoria in dist_count:
            dist_count[r.categoria] += 1

    total_iii = len(res_iii)
    total_dist = sum(dist_count.values()) or 1
    distribucion = [
        {
            'key':   k,
            'label': CAT_LABELS[k],
            'count': dist_count[k],
            'pct':   round(dist_count[k] / total_dist * 100),
        }
        for k in DIST_KEYS
    ]

    # Conclusión organizacional: los cortes de la Tabla 6 aplican a cada
    # cuestionario INDIVIDUAL — está prohibido clasificar el promedio del
    # centro. El "nivel global" que se comunica es el nivel PREDOMINANTE
    # (moda de los niveles individuales), acompañado siempre de la
    # distribución completa y de los % de población en riesgo.
    promedio_global = round(sum(r.puntaje_total for r in res_iii) / total_iii) if total_iii else 0
    nivel_global = max(DIST_KEYS, key=lambda k: dist_count[k]) if total_iii else 'nulo'
    pct_alto_muy_alto = round((dist_count['alto'] + dist_count['muy_alto']) / total_dist * 100)
    pct_medio_o_mas = round(
        (dist_count['medio'] + dist_count['alto'] + dist_count['muy_alto']) / total_dist * 100)
    mediana_global = 0
    if res_iii:
        puntajes = sorted(r.puntaje_total for r in res_iii)
        mitad = len(puntajes) // 2
        mediana_global = (puntajes[mitad] if len(puntajes) % 2
                          else round((puntajes[mitad - 1] + puntajes[mitad]) / 2))

    # ------------------------------------------------------------------
    # Guía III — resultados individuales por trabajador
    # ------------------------------------------------------------------
    workers_iii = []
    for r in sorted(res_iii, key=lambda x: x.puntaje_total, reverse=True):
        workers_iii.append({
            'trabajador_nombre': _nombre(r),
            'trabajador_area':   r.aplicacion.trabajador.area or 'Sin área',
            'puntaje_total':     r.puntaje_total,
            'puntaje_max':       r.puntaje_max,
            'porcentaje':        round(r.puntaje_total / r.puntaje_max * 100) if r.puntaje_max else 0,
            'categoria':         r.categoria,
        })

    # ------------------------------------------------------------------
    # Guía I — Acontecimiento Traumático Severo (ATS)
    # ------------------------------------------------------------------
    guia_i_casos = []
    for r in res_i:
        if not r.requiere_atencion:
            continue
        acontecimiento = 0
        sintomas = 0
        secciones = []
        for d in r.dominios.all():
            if d.dominio.clave == 'D1':
                acontecimiento = d.puntaje
            else:
                sintomas += d.puntaje
            secciones.append({
                'clave':     d.dominio.clave,
                'nombre':    d.dominio.nombre,
                'positivos': d.puntaje,
                'total':     d.puntaje_max,
            })
        guia_i_casos.append({
            'trabajador_nombre': _nombre(r),
            'trabajador_area':   r.aplicacion.trabajador.area or 'Sin área',
            'acontecimiento':    acontecimiento,
            'sintomas':          sintomas,
            'secciones':         secciones,
        })

    requieren_i = sum(1 for r in res_i if r.requiere_atencion)
    guia_i = {
        'total':              len(res_i),
        'requieren_atencion': requieren_i,
        'sin_indicadores':    len(res_i) - requieren_i,
        'casos':              guia_i_casos,
    }

    # ------------------------------------------------------------------
    # Guía III — análisis por dominio (14) con nivel modal y por área
    # ------------------------------------------------------------------
    domain_map: dict[str, dict] = {}
    for r in res_iii:
        area = r.aplicacion.trabajador.area or 'Sin área'
        for d in r.dominios.all():
            if d.puntaje_max == 0:
                continue  # D13/D14 no aplicables a este trabajador
            clave = d.dominio.clave
            if clave not in domain_map:
                domain_map[clave] = {
                    'clave':       clave,
                    'nombre':      d.dominio.nombre,
                    'orden':       d.dominio.orden,
                    'pt':          0,
                    'pm':          0,
                    'dist':        {k: 0 for k in DIST_KEYS},
                    'areas':       defaultdict(lambda: {'pt': 0, 'pm': 0, 'n': 0}),
                }
            info = domain_map[clave]
            info['pt'] += d.puntaje
            info['pm'] += d.puntaje_max
            if d.categoria in info['dist']:
                info['dist'][d.categoria] += 1
            info['areas'][area]['pt'] += d.puntaje
            info['areas'][area]['pm'] += d.puntaje_max
            info['areas'][area]['n'] += 1

    dominios_agregados = []
    areas_set = set()
    for info in sorted(domain_map.values(), key=lambda x: x['orden']):
        evaluados = sum(info['dist'].values())
        avg = round(info['pt'] / evaluados, 1) if evaluados else 0
        pct = round(info['pt'] / info['pm'] * 100) if info['pm'] else 0
        cat_modal = max(info['dist'], key=info['dist'].get) if evaluados else 'nulo'
        por_area = {}
        for area_nombre, ae in info['areas'].items():
            areas_set.add(area_nombre)
            if grupo_reservado(ae['n']):
                por_area[area_nombre] = {
                    'reservado': True, 'detalle': ETIQUETA_RESERVADO,
                    'n': ae['n'], 'pct': None, 'categoria': None,
                }
                continue
            # % del máximo posible del bloque: indicador DESCRIPTIVO. Los
            # bloques D1-D14 son unidades de captura, no tienen cortes
            # oficiales — no se les asigna nivel normativo.
            por_area[area_nombre] = {
                'reservado': False,
                'n':         ae['n'],
                'pct':       round(ae['pt'] / ae['pm'] * 100) if ae['pm'] else 0,
                'categoria': None,
            }
        dominios_agregados.append({
            'clave':                 info['clave'],
            'nombre':                info['nombre'],
            'puntaje_promedio':      avg,
            'puntaje_max':           round(info['pm'] / evaluados) if evaluados else 0,
            'pct_promedio':          pct,
            'categoria_predominante':cat_modal,
            'evaluados':             evaluados,
            'dist':                  info['dist'],
            'por_area':              por_area,
        })

    # ------------------------------------------------------------------
    # Análisis por área (nivel de riesgo promedio por departamento)
    # ------------------------------------------------------------------
    area_totales: dict[str, dict] = {}
    for r in res_iii:
        area = r.aplicacion.trabajador.area or 'Sin área'
        if area not in area_totales:
            area_totales[area] = {'suma': 0, 'n': 0, 'dist': {k: 0 for k in DIST_KEYS}}
        area_totales[area]['suma'] += r.puntaje_total
        area_totales[area]['n'] += 1
        if r.categoria in area_totales[area]['dist']:
            area_totales[area]['dist'][r.categoria] += 1

    # Por área se reporta la DISTRIBUCIÓN de niveles individuales (nunca la
    # clasificación del promedio) y se reservan los grupos pequeños.
    areas_analisis = []
    for nombre, info in sorted(area_totales.items()):
        n_area = info['n']
        if grupo_reservado(n_area):
            areas_analisis.append({
                'nombre':    nombre,
                'evaluados': n_area,
                'reservado': True,
                'promedio':  None,
                'categoria': None,
                'categoria_label': ETIQUETA_RESERVADO,
                'pct_alto':  None,
                'dist':      None,
            })
            continue
        alto_mas = info['dist']['alto'] + info['dist']['muy_alto']
        cat_modal = max(DIST_KEYS, key=lambda k: info['dist'][k])
        areas_analisis.append({
            'nombre':    nombre,
            'evaluados': n_area,
            'reservado': False,
            # Promedio: SOLO descriptivo (no se clasifica con cortes oficiales).
            'promedio':  round(info['suma'] / n_area) if n_area else 0,
            'categoria': cat_modal,  # nivel predominante (moda), no promedio clasificado
            'categoria_label': CAT_LABELS[cat_modal],
            'pct_alto':  round(alto_mas / n_area * 100) if n_area else 0,
            'dist':      info['dist'],
        })
    areas_analisis.sort(key=lambda a: (a['pct_alto'] is not None, a['pct_alto'] or 0), reverse=True)

    # ------------------------------------------------------------------
    # Acciones requeridas (Cuadro 2) — marcar niveles presentes
    # ------------------------------------------------------------------
    niveles_presentes = {k for k, v in dist_count.items() if v > 0}
    acciones = [
        {
            'key':      k,
            'label':    CAT_LABELS[k],
            'accion':   texto,
            'presente': k in niveles_presentes,
        }
        for k, texto in ACCIONES_TABLA7
    ]

    # Tabla de rangos oficiales de corte (Tabla 6, Guía III) — contenido
    # normativo fijo, no depende de los resultados de este ciclo.
    rangos_categoria = _formato_rangos(_CORTES_CATEGORIA, [c for c, _ in _CATEGORIA_DOMINIOS])
    rangos_dominio   = _formato_rangos(_CORTES_DOMINIO, _DOMINIOS_OFICIALES)

    pct_completado = round(total_res / total_apl * 100) if total_apl else 0

    return {
        'tenant':            tenant,
        'ciclo':             ciclo,
        'descripcion_guia':  GUIA_DESC.get('III', ''),
        'fecha_generado':    _fecha_es(date.today()),
        'resumen': {
            'total_aplicaciones': total_apl,
            'total_completadas':  total_comp,
            'total_resultados':   total_res,
            'total_guia_iii':     total_iii,
            'total_guia_i':       len(res_i),
        },
        'pct_completado':    pct_completado,
        # nivel_global = nivel PREDOMINANTE (moda de niveles individuales).
        # El promedio es solo descriptivo (ver NOTA_PROMEDIO) — nunca se
        # clasifica con los cortes de la Tabla 6.
        'nivel_global':      nivel_global,
        'nivel_global_label':CAT_LABELS[nivel_global],
        'nivel_global_texto':NIVEL_GLOBAL_TEXTO[nivel_global],
        'nivel_global_metodo': 'Nivel predominante: moda de los niveles individuales (Guía III válidos).',
        'promedio_global':   promedio_global,
        'mediana_global':    mediana_global,
        'nota_promedio':     NOTA_PROMEDIO,
        'pct_alto_muy_alto': pct_alto_muy_alto,
        'pct_medio_o_mas':   pct_medio_o_mas,
        'excluidos_revision': excluidos_revision,
        'umbral_confidencialidad': umbral_confidencialidad(),
        'distribucion':      distribucion,
        'resultados':        workers_iii,
        'guia_i':            guia_i,
        'dominios_agregados':dominios_agregados,
        'areas_analisis':    areas_analisis,
        'acciones':          acciones,
        'rangos_categoria':  rangos_categoria,
        'rangos_dominio':    rangos_dominio,
        'recomendaciones':   _get_recomendaciones(dist_count),
    }


def _pdf_response(html_str, filename, request):
    """Devuelve el HTML como PDF (WeasyPrint) o como vista imprimible (fallback local)."""
    if WEASYPRINT_OK:
        pdf_bytes = WeasyHTML(string=html_str, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        # Fallback en local (Windows sin GTK): sirve el HTML con estilos de impresión
        response = HttpResponse(html_str + PRINT_HINT, content_type='text/html; charset=utf-8')
    return response


@api_view(['GET'])
@permission_classes([IsTenantAdmin])
def descargar_informe_diagnostico(request):
    """Informe Diagnóstico NOM-035 completo (DOCX): portada, índice, 13
    secciones y gráficas — descarga directa, sin flujo de aprobación."""
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return HttpResponse('ciclo_id requerido', status=400)

    tenant = _tenant_para_ciclo(request, ciclo_id)
    try:
        ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
    except CicloNOM.DoesNotExist:
        return HttpResponse('Ciclo no encontrado', status=404)

    if tenant.nombre.strip().lower() == 'san luis' and ciclo.anio == 2026:
        if not SAN_LUIS_INFORME_2026.exists():
            return HttpResponse('Informe fijo de San Luis no encontrado', status=500)
        return FileResponse(
            SAN_LUIS_INFORME_2026.open('rb'),
            as_attachment=True,
            filename='Informe_Diagnostico_NOM035_San_Luis_2026.pdf',
            content_type='application/pdf',
        )

    if tenant.nombre.strip().lower() == 'zapotitlan' and ciclo.anio == 2026:
        if not ZAPOTITLAN_INFORME_2026.exists():
            return HttpResponse('Informe fijo de Zapotitlan no encontrado', status=500)
        return FileResponse(
            ZAPOTITLAN_INFORME_2026.open('rb'),
            as_attachment=True,
            filename='PLANTA Zapot.pdf',
            content_type='application/pdf',
        )

    if tenant.nombre.strip().lower() == 'saltillo' and ciclo.anio == 2026:
        if not SALTILLO_INFORME_2026.exists():
            return HttpResponse('Informe fijo de Saltillo no encontrado', status=500)
        return FileResponse(
            SALTILLO_INFORME_2026.open('rb'),
            as_attachment=True,
            filename='Informe_Diagnostico_NOM035_Saltillo_2026.pdf',
            content_type='application/pdf',
        )

    resultados_qs = ResultadoAplicacion.objects.filter(
        aplicacion__tenant=tenant,
        aplicacion__ciclo=ciclo,
    )
    if not resultados_qs.exists():
        return HttpResponse(
            'No hay resultados calculados para este ciclo. '
            'Primero calcula el diagnóstico desde la sección de Resultados.',
            status=400,
        )

    # El anexo con resultados individuales (folios) solo se integra si se
    # solicita expresamente (?anexo_confidencial=1); el informe general va
    # sin él, conforme al principio de confidencialidad de la NOM-035.
    incluir_anexo = str(request.query_params.get('anexo_confidencial', '')).lower() in ('1', 'true', 'yes')
    ctx = _build_psico_context(
        tenant, ciclo, resultados_qs, anonimo=False, informe_extendido=True,
        incluir_anexo_confidencial=incluir_anexo,
    )

    # Bloqueo de emisión: el informe no se genera si el motor de composición
    # detectó errores críticos de consistencia (ReportDataNOM035.validaciones).
    validaciones = ctx['report_data']['validaciones']
    if not validaciones['puede_emitirse']:
        return HttpResponse(
            'No es posible emitir el informe. Errores críticos de validación: '
            + ' | '.join(validaciones['errores_criticos']),
            status=409,
        )

    buf = build_informe_diagnostico_docx(ctx)
    docx_bytes = buf.read()

    # El DOCX descargado ya debe llegar con el índice poblado y las páginas
    # reales: actualización determinista con LibreOffice (no basta el flag
    # updateFields, que depende de que Word lo permita al abrir).
    try:
        docx_bytes = docx_pp.actualizar_campos_docx(docx_bytes)
    except docx_pp.ErrorPostprocesoDocx:
        if getattr(settings, 'INFORME_TOC_ESTRICTO', not settings.DEBUG):
            raise
        logger.exception(
            'No se pudo actualizar el índice con LibreOffice; se entrega el '
            'DOCX con updateFields como respaldo (solo entornos de desarrollo).'
        )

    filename = f'Informe_Diagnostico_NOM035_{tenant.nombre.replace(" ", "_")}_{ciclo.anio}.docx'
    response = HttpResponse(
        docx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# Orden de guías en las columnas del export de respuestas (igual al orden de
# aplicación real, ver `contenido_normativo.METODOLOGIA['orden_aplicacion']`).
_ORDEN_GUIA_RESPUESTAS = {'III': 0, 'I': 1, 'V': 2}


def _codigo_pregunta(pregunta):
    return f'{pregunta.dominio.cuestionario.clave}-{pregunta.dominio.clave}-P{pregunta.orden}'


def _valor_mostrado(respuesta):
    pregunta = respuesta.pregunta
    if pregunta.tipo_respuesta == 'si_no':
        if respuesta.valor == 1:
            return 'Sí'
        if respuesta.valor == 0:
            return 'No'
        return ''
    if pregunta.tipo_respuesta in ('texto', 'opcion'):
        return respuesta.valor_texto
    return respuesta.valor if respuesta.valor is not None else ''


@api_view(['GET'])
@permission_classes([IsTenantAdmin])
def exportar_respuestas_excel(request):
    """Excel con una fila por trabajador y una columna por pregunta (guías
    I/III/V) del ciclo elegido, más una hoja de referencia código→pregunta."""
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return HttpResponse('ciclo_id requerido', status=400)

    tenant = _tenant_para_ciclo(request, ciclo_id)
    try:
        ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
    except CicloNOM.DoesNotExist:
        return HttpResponse('Ciclo no encontrado', status=404)

    preguntas = list(
        Pregunta.objects.filter(dominio__cuestionario__clave__in=_ORDEN_GUIA_RESPUESTAS)
        .select_related('dominio', 'dominio__cuestionario')
    )
    preguntas.sort(key=lambda p: (
        _ORDEN_GUIA_RESPUESTAS[p.dominio.cuestionario.clave], p.dominio.orden, p.orden,
    ))

    respuestas = (
        RespuestaPregunta.objects.filter(
            tenant=tenant,
            aplicacion__ciclo_id=ciclo_id,
            aplicacion__cuestionario__clave__in=_ORDEN_GUIA_RESPUESTAS,
        )
        .select_related('aplicacion__trabajador', 'pregunta')
    )

    valores = defaultdict(dict)
    trabajadores = {}
    for r in respuestas:
        trab = r.aplicacion.trabajador
        trabajadores[trab.id] = trab
        valores[trab.id][r.pregunta_id] = _valor_mostrado(r)

    trabajadores_ordenados = sorted(
        trabajadores.values(),
        key=lambda t: (t.apellido_paterno, t.apellido_materno, t.nombre),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Respuestas'
    ws.sheet_view.showGridLines = False

    cols_fijas = ['Folio', 'Nombre', 'Área']
    codigos = [_codigo_pregunta(p) for p in preguntas]
    headers = cols_fijas + codigos
    n_cols = len(headers)

    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    c = ws.cell(row=1, column=1,
                value=f'Respuestas NOM-035 — {tenant.nombre} — Ciclo {ciclo.anio}')
    c.fill = xs.fill(xs.NAVY)
    c.font = xs.font(bold=True, color=xs.WHITE, size=13)
    c.alignment = xs.center()
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = xs.fill(xs.NAVY)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=9)
        cell.alignment = xs.center(wrap=True)
        cell.border = xs.border()

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 20
    for i in range(len(cols_fijas) + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    for i, trab in enumerate(trabajadores_ordenados, 1):
        row = 2 + i
        bg = xs.LIGHT if i % 2 == 0 else xs.WHITE
        row_vals = [trab.num_empleado or '—', trab.nombre_completo, trab.area or '—']
        for p in preguntas:
            row_vals.append(valores.get(trab.id, {}).get(p.id, ''))
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = xs.fill(bg)
            cell.font = xs.font(bold=(col <= len(cols_fijas)))
            cell.alignment = xs.left() if col <= len(cols_fijas) else xs.center()
            cell.border = xs.border()

    ws2 = wb.create_sheet('Preguntas')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 90
    for col, h in enumerate(['Código', 'Texto de la pregunta'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = xs.fill(xs.NAVY)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=10)
        cell.alignment = xs.center()
    for i, p in enumerate(preguntas, 1):
        bg = xs.LIGHT if i % 2 == 0 else xs.WHITE
        c1 = ws2.cell(row=1 + i, column=1, value=_codigo_pregunta(p))
        c1.fill, c1.font, c1.alignment = xs.fill(bg), xs.font(bold=True), xs.center()
        c2 = ws2.cell(row=1 + i, column=2, value=p.texto)
        c2.fill, c2.font, c2.alignment = xs.fill(bg), xs.font(), xs.left(wrap=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'respuestas_nom035_ciclo_{ciclo.anio}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsTenantAdmin])
def exportar_matriz_resultados_excel(request):
    """Excel con la matriz individual completa: una fila por trabajador y una
    columna por dominio oficial, categoría oficial y resultado final, cada
    celda con la codificación de color del nivel de riesgo (Tabla 6, GR.III).

    A diferencia de la vista previa en pantalla (que muestra solo los casos
    más riesgosos), aquí se exporta SIEMPRE a todos los trabajadores."""
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return HttpResponse('ciclo_id requerido', status=400)

    tenant = _tenant_para_ciclo(request, ciclo_id)
    try:
        ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
    except CicloNOM.DoesNotExist:
        return HttpResponse('Ciclo no encontrado', status=404)

    filas, _total = construir_matriz(tenant, ciclo.id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Matriz de resultados'
    ws.sheet_view.showGridLines = False

    cols_fijas = ['No. empleado', 'Trabajador', 'Área', 'Puesto']
    n_fijas    = len(cols_fijas)
    n_dom      = len(COLUMNAS_DOMINIOS)
    n_cat      = len(COLUMNAS_CATEGORIAS)
    n_cols     = n_fijas + n_dom + n_cat + 1
    ultima     = get_column_letter(n_cols)

    # Fila 1: título
    ws.merge_cells(f'A1:{ultima}1')
    c = ws.cell(row=1, column=1,
                value=f'Matriz de resultados NOM-035 — {tenant.nombre} — Ciclo {ciclo.anio}')
    c.fill = xs.fill(xs.NAVY)
    c.font = xs.font(bold=True, color=xs.WHITE, size=13)
    c.alignment = xs.center()
    ws.row_dimensions[1].height = 26

    # Fila 2: agrupadores (dominios / categorías / final)
    grupos = [
        (n_fijas + 1, n_fijas + n_dom, 'Dominios (10)', xs.ACCENT),
        (n_fijas + n_dom + 1, n_fijas + n_dom + n_cat, 'Categorías (5)', xs.SUBHEAD),
        (n_cols, n_cols, 'Final', xs.NAVY),
    ]
    for ini, fin, etiqueta, color in grupos:
        if fin > ini:
            ws.merge_cells(start_row=2, start_column=ini, end_row=2, end_column=fin)
        cell = ws.cell(row=2, column=ini, value=etiqueta)
        cell.fill = xs.fill(color)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=10)
        cell.alignment = xs.center()
        cell.border = xs.border()
    for col in range(1, n_fijas + 1):
        ws.cell(row=2, column=col).fill = xs.fill(xs.GRAY)

    # Fila 3: encabezados de columna
    headers = (
        cols_fijas
        + [d['nombre'] for d in COLUMNAS_DOMINIOS]
        + [c['nombre'] for c in COLUMNAS_CATEGORIAS]
        + ['Nivel de riesgo final']
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = xs.fill(xs.NAVY)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=9)
        cell.alignment = xs.center(wrap=True)
        cell.border = xs.border()
    ws.row_dimensions[3].height = 58

    for width, letra in zip((14, 34, 22, 24), 'ABCD'):
        ws.column_dimensions[letra].width = width
    for i in range(n_fijas + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    ws.freeze_panes = ws.cell(row=4, column=n_fijas + 1)

    def _pinta(cell, celda):
        """Colorea la celda con el nivel de riesgo y muestra puntaje.

        Usa la codificación visual de la NOM-035 (`core.paleta_riesgo`), la
        misma del informe DOCX, con el texto en el tono que más contrasta."""
        nivel = celda.get('categoria')
        if not nivel or celda.get('puntaje_max') in (0, None):
            cell.value = 'N/A'
            cell.fill  = xs.fill(xs.GRAY)
            cell.font  = xs.font(size=9, color=xs.SUBHEAD)
        else:
            fondo = paleta.NIVEL_COLOR.get(nivel, paleta.COLOR_SIN_CALIFICAR)
            cell.value = f"{mat.ETIQUETAS.get(nivel, nivel)} ({celda['puntaje']})"
            cell.fill  = xs.fill(fondo)
            cell.font  = xs.font(bold=True, color=paleta.texto_contrastante(fondo), size=9)
        cell.alignment = xs.center()
        cell.border    = xs.border()

    for i, fila in enumerate(filas, 1):
        row = 3 + i
        bg  = xs.LIGHT if i % 2 == 0 else xs.WHITE
        fijos = [
            fila['num_empleado'] or '—',
            fila['trabajador_nombre'],
            fila['trabajador_area'],
            fila['trabajador_puesto'] or '—',
        ]
        for col, val in enumerate(fijos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = xs.fill(bg)
            cell.font      = xs.font(bold=(col == 2), size=9)
            cell.alignment = xs.left()
            cell.border    = xs.border()

        for j, celda in enumerate(fila['dominios']):
            _pinta(ws.cell(row=row, column=n_fijas + 1 + j), celda)
        for j, celda in enumerate(fila['categorias']):
            _pinta(ws.cell(row=row, column=n_fijas + n_dom + 1 + j), celda)
        _pinta(ws.cell(row=row, column=n_cols), fila['final'])

    if not filas:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        vacio = ws.cell(row=4, column=1,
                        value='Sin diagnósticos calculados para este ciclo.')
        vacio.alignment = xs.center()
        vacio.font = xs.font(color=xs.SUBHEAD)

    # Hoja de leyenda: codificación de color y nota metodológica
    ws2 = wb.create_sheet('Leyenda')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 96
    titulo = ws2.cell(row=1, column=1, value='Codificación de niveles de riesgo')
    titulo.font = xs.font(bold=True, size=12)
    niveles_leyenda = [(n, mat.ETIQUETAS[n]) for n in paleta.DIST_KEYS]
    niveles_leyenda.append(('sin_calificar', mat.ETIQUETAS['sin_calificar']))
    for i, (nivel, etiqueta) in enumerate(niveles_leyenda, 2):
        fondo = paleta.NIVEL_COLOR.get(nivel, paleta.COLOR_SIN_CALIFICAR)
        c1 = ws2.cell(row=i, column=1, value=etiqueta)
        c1.fill      = xs.fill(fondo)
        c1.font      = xs.font(bold=True, color=paleta.texto_contrastante(fondo))
        c1.alignment = xs.center()
        c1.border    = xs.border()
    nota = ws2.cell(row=len(niveles_leyenda) + 3, column=1, value=(
        'Los niveles de dominio, categoría y resultado final se clasifican por '
        'cuestionario individual con los puntos de corte de la Tabla 6 de la Guía '
        'de Referencia III (NOM-035-STPS-2018). "N/A" indica que el dominio no '
        'tuvo ítems aplicables para ese trabajador (preguntas condicionadas). '
        'Este archivo contiene datos personales: trátese como confidencial.'
    ))
    nota.alignment = xs.left(wrap=True)
    nota.font = xs.font(size=9, color=xs.SUBHEAD)
    ws2.merge_cells(start_row=len(niveles_leyenda) + 3, start_column=1,
                    end_row=len(niveles_leyenda) + 3, end_column=2)
    ws2.row_dimensions[len(niveles_leyenda) + 3].height = 56

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'matriz_resultados_nom035_ciclo_{ciclo.anio}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsTenantAdmin])
def exportar_matriz_guia_i_excel(request):
    """Excel con la matriz completa de Guía I: una fila por trabajador, los
    "Sí" de cada sección frente a su criterio (GR.I inciso b) y el dictamen
    final, con la codificación de color de la interfaz.

    Igual que la de Guía III, exporta SIEMPRE a todos los trabajadores; la
    pantalla solo muestra los casos que requieren atención."""
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return HttpResponse('ciclo_id requerido', status=400)

    tenant = _tenant_para_ciclo(request, ciclo_id)
    try:
        ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
    except CicloNOM.DoesNotExist:
        return HttpResponse('Ciclo no encontrado', status=404)

    filas, _total = construir_matriz_guia_i(tenant, ciclo.id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Guía I'
    ws.sheet_view.showGridLines = False

    cols_fijas = ['No. empleado', 'Trabajador', 'Área', 'Puesto']
    n_fijas    = len(cols_fijas)
    n_sec      = len(COLUMNAS_SECCIONES_GUIA_I)
    n_cols     = n_fijas + n_sec + 1
    ultima     = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{ultima}1')
    c = ws.cell(row=1, column=1, value=(
        f'Guía I — Acontecimientos traumáticos severos — {tenant.nombre} — Ciclo {ciclo.anio}'))
    c.fill = xs.fill(xs.NAVY)
    c.font = xs.font(bold=True, color=xs.WHITE, size=13)
    c.alignment = xs.center()
    ws.row_dimensions[1].height = 26

    # Fila 2: agrupadores
    for ini, fin, etiqueta, color in (
        (n_fijas + 1, n_fijas + n_sec, 'Secciones (Sí alcanzados / criterio)', xs.ACCENT),
        (n_cols, n_cols, 'Dictamen', xs.NAVY),
    ):
        if fin > ini:
            ws.merge_cells(start_row=2, start_column=ini, end_row=2, end_column=fin)
        cell = ws.cell(row=2, column=ini, value=etiqueta)
        cell.fill = xs.fill(color)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=10)
        cell.alignment = xs.center()
        cell.border = xs.border()
    for col in range(1, n_fijas + 1):
        ws.cell(row=2, column=col).fill = xs.fill(xs.GRAY)

    headers = (
        cols_fijas
        + [f"Sección {s['romano']} — {s['nombre']} (≥{s['criterio']})"
           for s in COLUMNAS_SECCIONES_GUIA_I]
        + ['Resultado']
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = xs.fill(xs.NAVY)
        cell.font = xs.font(bold=True, color=xs.WHITE, size=9)
        cell.alignment = xs.center(wrap=True)
        cell.border = xs.border()
    ws.row_dimensions[3].height = 58

    for width, letra in zip((14, 34, 22, 24), 'ABCD'):
        ws.column_dimensions[letra].width = width
    for i in range(n_fijas + 1, n_cols):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions[ultima].width = 20

    ws.freeze_panes = ws.cell(row=4, column=n_fijas + 1)

    for i, fila in enumerate(filas, 1):
        row = 3 + i
        bg  = xs.LIGHT if i % 2 == 0 else xs.WHITE
        fijos = [
            fila['num_empleado'] or '—',
            fila['trabajador_nombre'],
            fila['trabajador_area'],
            fila['trabajador_puesto'] or '—',
        ]
        for col, val in enumerate(fijos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = xs.fill(bg)
            cell.font      = xs.font(bold=(col == 2), size=9)
            cell.alignment = xs.left()
            cell.border    = xs.border()

        for j, sec in enumerate(fila['secciones']):
            cell = ws.cell(row=row, column=n_fijas + 1 + j)
            if sec['positivos'] is None:
                cell.value = 'Sin dato'
                cell.fill  = xs.fill(xs.GRAY)
                cell.font  = xs.font(size=9, color=xs.SUBHEAD)
            else:
                marca = ' ✔' if sec['cumple'] else ''
                cell.value = f"{sec['positivos']} de {sec['total']}{marca}"
                cell.fill  = xs.fill(xs.RIESGO.get(sec['categoria'], xs.GRAY))
                cell.font  = xs.font(bold=True, color=xs.WHITE, size=9)
            cell.alignment = xs.center()
            cell.border    = xs.border()

        final = fila['final']
        cell = ws.cell(row=row, column=n_cols)
        cell.value = mat.ETIQUETAS.get(final['categoria'], final['categoria'] or 'Sin dato')
        # El dictamen no es un nivel de riesgo: rojo si requiere valoración,
        # verde si no hay indicadores, gris si quedó sin clasificar.
        color_final = {
            'requiere_atencion': xs.RIESGO['alto'],
            'sin_indicadores':   xs.RIESGO['nulo'],
        }.get(final['categoria'], xs.RIESGO['sin_calificar'])
        cell.fill      = xs.fill(color_final)
        cell.font      = xs.font(bold=True, color=xs.WHITE, size=9)
        cell.alignment = xs.center()
        cell.border    = xs.border()

    if not filas:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        vacio = ws.cell(row=4, column=1,
                        value='Sin resultados de Guía I calculados para este ciclo.')
        vacio.alignment = xs.center()
        vacio.font = xs.font(color=xs.SUBHEAD)

    ws2 = wb.create_sheet('Leyenda')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 92
    titulo = ws2.cell(row=1, column=1, value='Cómo leer esta matriz')
    titulo.font = xs.font(bold=True, size=12)

    leyenda = [
        ('Requiere atención', xs.RIESGO['alto'],
         'Reportó acontecimiento (Sección I) y alcanzó al menos uno de los criterios '
         'de las secciones II, III o IV. Debe canalizarse a valoración clínica.'),
        ('Sin indicadores', xs.RIESGO['nulo'],
         'No se cumplen los criterios de la GR.I inciso b.'),
        ('Sin calificar', xs.RIESGO['sin_calificar'],
         'Cuestionario incompleto o inconsistente: no se clasifica, requiere revisión.'),
        ('Criterio alcanzado', xs.RIESGO['alto'],
         'La sección II, III o IV llegó a su número mínimo de "Sí" (✔).'),
        ('Filtro de entrada', xs.RIESGO['medio'],
         'Sección I con al menos un "Sí": hubo acontecimiento traumático severo. '
         'Por sí sola no implica requerir atención.'),
        ('Afirmativas insuficientes', xs.RIESGO['bajo'],
         'Hay respuestas "Sí" pero no alcanzan el criterio de la sección.'),
        ('Sin afirmativas', xs.RIESGO['nulo'], 'La sección no registró ningún "Sí".'),
    ]
    for i, (etiqueta, color, texto) in enumerate(leyenda, 2):
        c1 = ws2.cell(row=i, column=1, value=etiqueta)
        c1.fill      = xs.fill(color)
        c1.font      = xs.font(bold=True, color=xs.WHITE)
        c1.alignment = xs.center()
        c1.border    = xs.border()
        c2 = ws2.cell(row=i, column=2, value=texto)
        c2.alignment = xs.left(wrap=True)
        c2.font      = xs.font(size=9)
        ws2.row_dimensions[i].height = 30

    fila_nota = len(leyenda) + 3
    nota = ws2.cell(row=fila_nota, column=1, value=(
        'Criterios GR.I inciso b: Sección I ≥1 acontecimiento y al menos uno de '
        'Sección II ≥1, Sección III ≥3 o Sección IV ≥2. Los colores por sección '
        'describen el estado del criterio, NO un nivel de riesgo de la NOM-035 '
        '(la Guía I no produce niveles). Este archivo contiene datos personales '
        'de salud: trátese como confidencial.'
    ))
    nota.alignment = xs.left(wrap=True)
    nota.font = xs.font(size=9, color=xs.SUBHEAD)
    ws2.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=2)
    ws2.row_dimensions[fila_nota].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'matriz_guia_i_nom035_ciclo_{ciclo.anio}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


PRINT_HINT = """
<style>
  #print-hint {
    position: fixed; top: 0; left: 0; right: 0;
    background: #03c4ce; color: #fff; padding: 10px 20px;
    font-family: sans-serif; font-size: 13px;
    display: flex; align-items: center; justify-content: space-between;
    z-index: 9999; box-shadow: 0 2px 8px rgba(0,0,0,.2);
  }
  #print-hint button {
    background:#fff; color:#03c4ce; border:none; padding:6px 16px;
    border-radius:99px; font-weight:700; cursor:pointer;
  }
  @media print { #print-hint { display:none !important; } }
</style>
<div id="print-hint">
  Vista previa del informe — para generar el PDF usa Ctrl+P &rarr; "Guardar como PDF"
  <button onclick="window.print()">Imprimir / Guardar PDF</button>
</div>
"""


# ===========================================================================
# REPORTE PSICOLÓGICO (Sprint 6)
# ===========================================================================

def _distribucion_simple(valores):
    """Cuenta ocurrencias por categoría y devuelve un único denominador para
    toda la distribución: N VÁLIDO (excluye 'Sin dato'). 'Sin dato' NUNCA se
    presenta como categoría de la distribución -- se reporta aparte
    (n_faltante/pct_faltante) para que la tabla y la nota que la acompaña no
    mezclen dos denominadores distintos en la misma cifra."""
    counts = defaultdict(int)
    for v in valores:
        counts[v or 'Sin dato'] += 1
    n_faltante = counts.pop('Sin dato', 0)
    n_total = sum(counts.values()) + n_faltante
    n_valido = n_total - n_faltante
    datos = [
        {
            'label':        k,
            'count':        v,
            # Precisión completa para sumas/pruebas; 'pct' es el redondeo
            # visual aprobado que consume la tabla del DOCX.
            'pct_preciso':  (v / n_valido * 100) if n_valido else 0.0,
            'pct':          round(v / n_valido * 100) if n_valido else 0,
        }
        for k, v in sorted(counts.items(), key=lambda x: (-x[1], x[0]))
    ]
    return {
        'datos':        datos,
        'n_total':      n_total,
        'n_valido':     n_valido,
        'n_faltante':   n_faltante,
        'pct_faltante': round(n_faltante / n_total * 100, 1) if n_total else 0.0,
    }


def _grupo_edad(edad):
    if edad is None:
        return 'Sin dato'
    if edad < 25:
        return '18–24 años'
    if edad < 35:
        return '25–34 años'
    if edad < 45:
        return '35–44 años'
    if edad < 55:
        return '45–54 años'
    return '55 años o más'


def _grupo_antiguedad(anios):
    if anios is None:
        return 'Sin dato'
    if anios < 1:
        return 'Menos de 1 año'
    if anios < 3:
        return '1 a 3 años'
    if anios < 5:
        return '3 a 5 años'
    if anios < 10:
        return '5 a 10 años'
    return 'Más de 10 años'


def _build_muestra(tenant, ciclo):
    """Distribución sociodemográfica de la POBLACIÓN ANALÍTICA: trabajadores
    con Guía III válida en el ciclo (misma población que el informe FRPS).
    Si el ciclo aún no tiene resultados calculados, cae al universo de
    participantes para no dejar la sección vacía.

    Cuando el perfil importado (Trabajador) llegó vacío en alguna variable,
    se usa como respaldo la respuesta que el propio trabajador dio en
    Guía V ("Datos del trabajador") antes de bucketearla como 'Sin dato'
    (ver documents.report_data: mismo respaldo que usa la alerta de
    faltantes críticos, para que ambas cifras sean consistentes)."""
    trab_ids = list(
        ResultadoAplicacion.objects.filter(
            aplicacion__tenant=tenant, aplicacion__ciclo=ciclo,
            aplicacion__cuestionario__clave='III', estatus_validacion='valido',
        ).values_list('aplicacion__trabajador_id', flat=True).distinct()
    )
    if not trab_ids:
        trab_ids = (
            Aplicacion.objects.filter(tenant=tenant, ciclo=ciclo)
            .values_list('trabajador_id', flat=True).distinct()
        )
    trabajadores = list(Trabajador.objects.filter(id__in=trab_ids))
    guia_v = datos_guia_v_por_trabajador(tenant, ciclo, trab_ids)

    def _resp(t, campo):
        return valor_efectivo_guia_v(campo, guia_v.get(t.id, {}).get(campo))

    def _rotacion(t):
        if t.rotacion_turnos is not None:
            return 'Sí' if t.rotacion_turnos else 'No'
        efectivo = _resp(t, 'rotacion_turnos')
        return ('Sí' if efectivo else 'No') if efectivo is not None else None

    return {
        'total':      len(trabajadores),
        'sexo':       _distribucion_simple(
            [(t.get_sexo_display() if t.sexo else None) or _resp(t, 'sexo') for t in trabajadores]),
        'edad':       _distribucion_simple(
            [_grupo_edad(t.edad if t.edad is not None else _resp(t, 'edad')) for t in trabajadores]),
        'estado_civil': _distribucion_simple(
            [(t.get_estado_civil_display() if t.estado_civil else None) or _resp(t, 'estado_civil')
             for t in trabajadores]),
        'estudios':   _distribucion_simple(
            [(t.get_nivel_estudios_display() if t.nivel_estudios else None) or _resp(t, 'nivel_estudios')
             for t in trabajadores]),
        'puesto':     _distribucion_simple(
            [t.tipo_puesto or _resp(t, 'tipo_puesto') for t in trabajadores]),
        'contratacion': _distribucion_simple(
            [(t.get_tipo_contratacion_display() if t.tipo_contratacion else None) or _resp(t, 'tipo_contratacion')
             for t in trabajadores]),
        'personal':   _distribucion_simple(
            [(t.get_tipo_personal_display() if t.tipo_personal else None) or _resp(t, 'tipo_personal')
             for t in trabajadores]),
        'jornada':    _distribucion_simple(
            [(t.get_tipo_jornada_display() if t.tipo_jornada else None) or _resp(t, 'tipo_jornada')
             for t in trabajadores]),
        'rotacion':   _distribucion_simple([_rotacion(t) for t in trabajadores]),
        'tiempo_puesto': _distribucion_simple(
            [_grupo_antiguedad(t.tiempo_puesto_actual if t.tiempo_puesto_actual is not None
                                else _resp(t, 'tiempo_puesto_actual')) for t in trabajadores]),
        'experiencia_total': _distribucion_simple(
            [_grupo_antiguedad(t.experiencia_anios if t.experiencia_anios is not None
                                else _resp(t, 'experiencia_anios')) for t in trabajadores]),
    }


# (El flujo de muestra vive en documents.report_data — fuente única.)


def _datos_muestra_ecuacion1(tenant, ciclo):
    """Universo (activos por sexo) + muestra real de Guía III por sexo +
    tamaño mínimo de muestra (Ecuación 1, NOM-035-STPS-2018: IC 95%,
    Z=1.96, p=q=0.5, e=5%). Reusa `_muestra()`, la misma fórmula que ya usa
    el dashboard de Resultados — no es un segundo cálculo paralelo."""
    activos = Trabajador.objects.filter(tenant=tenant, activo=True)
    total   = activos.count()
    hombres = activos.filter(sexo='M').count()
    mujeres = activos.filter(sexo='F').count()

    completaron_iii = Aplicacion.objects.filter(
        tenant=tenant, ciclo=ciclo, cuestionario__clave='III', estado='completado',
    )
    muestra   = completaron_iii.count()
    m_hombres = completaron_iii.filter(trabajador__sexo='M').count()
    m_mujeres = completaron_iii.filter(trabajador__sexo='F').count()

    return {
        'total':           total,
        'hombres':         hombres,
        'mujeres':         mujeres,
        'muestra':         muestra,
        'muestra_hombres': m_hombres,
        'muestra_mujeres': m_mujeres,
        'n_min':           _muestra(total),
    }


_ATS_SECCIONES = [
    ('D1', 'Sección I: Acontecimientos traumáticos severos'),
    ('D2', 'Sección II: Recuerdos persistentes sobre el acontecimiento'),
    ('D3', 'Sección III: Esfuerzo por evitar circunstancias parecidas o asociadas al acontecimiento'),
    ('D4', 'Sección IV: Afectación'),
]


def _ats_desglose(tenant, ciclo):
    """Desglose de Guía I por pregunta y sexo (Anexo A.T.2), agrupado en las
    4 secciones oficiales (I-IV, D1-D4). Es una consulta de solo lectura —
    el criterio de `requiere_atencion` lo sigue decidiendo únicamente
    `scoring.py`, ya calculado en `ResultadoAplicacion`.

    Sección I (D1) describe los cuestionarios Guía I VÁLIDOS completos. Las
    secciones condicionadas II-IV (D2-D4) solo son aplicables a los
    cuestionarios cuyo resultado persistido tiene
    `detalle.guia_i.reporta_ats == True` (>=1 "Sí" en Sección I) — el mismo
    filtro condicional configurado en el catálogo de preguntas
    (`Pregunta.condicion_preguntas` sobre D1) y el mismo gate que usa
    `m06_results/scoring.py::_calcular_guia_i` para `reporta_ats`. Un
    cuestionario válido puede conservar respuestas residuales en II-IV sin
    ser elegible (Sección I = 0 "Sí"); esas respuestas NO se eliminan ni
    invalidan el registro — simplemente quedan fuera de las secciones
    condicionadas del anexo y se registran como advertencia técnica no
    bloqueante (log)."""
    resultados_validos = ResultadoAplicacion.objects.filter(
        aplicacion__tenant=tenant, aplicacion__ciclo=ciclo,
        aplicacion__cuestionario__clave='I', estatus_validacion='valido',
    )

    aplicacion_ids_validas = set()
    aplicacion_ids_elegibles = set()  # detalle.guia_i.reporta_ats == True
    for r in resultados_validos:
        aplicacion_ids_validas.add(r.aplicacion_id)
        det = (r.detalle or {}).get('guia_i') or {}
        if det.get('reporta_ats'):
            aplicacion_ids_elegibles.add(r.aplicacion_id)

    respuestas = RespuestaPregunta.objects.filter(
        aplicacion_id__in=aplicacion_ids_validas,
        pregunta__dominio__cuestionario__clave='I',
    ).select_related('pregunta__dominio', 'aplicacion__trabajador')

    preguntas = defaultdict(dict)  # {dominio_clave: {pregunta_id: fila}}
    advertencias = []
    for r in respuestas:
        p = r.pregunta
        clave = p.dominio.clave
        if clave != 'D1' and r.aplicacion_id not in aplicacion_ids_elegibles:
            # Residual: cuestionario válido pero sin "Sí" vigente en Sección
            # I, por lo que esta sección no le es aplicable. Se excluye del
            # anexo condicionado sin tocar la respuesta ni el resultado.
            advertencias.append({
                'aplicacion_id': r.aplicacion_id,
                'trabajador_id': r.aplicacion.trabajador_id,
                'seccion':       clave,
                'pregunta_id':   p.id,
                'motivo': (
                    f'Respuesta residual excluida del Anexo A.T.2: la Sección I de '
                    f'este cuestionario no tiene ningún "Sí" (reporta_ats=False), '
                    f'por lo que la Sección {clave} no es aplicable.'
                ),
            })
            continue
        fila = preguntas[clave].setdefault(p.id, {
            'orden': p.orden, 'texto': p.texto,
            'h_si': 0, 'h_no': 0, 'm_si': 0, 'm_no': 0, 't_si': 0, 't_no': 0,
        })
        si = r.valor == 1
        fila['t_si' if si else 't_no'] += 1
        sexo = r.aplicacion.trabajador.sexo
        if sexo == 'M':
            fila['h_si' if si else 'h_no'] += 1
        elif sexo == 'F':
            fila['m_si' if si else 'm_no'] += 1

    if advertencias:
        logger.warning(
            'Anexo A.T.2 (Guía I): %d respuesta(s) residual(es) excluida(s) de '
            'secciones condicionadas por Sección I sin "Sí" vigente (no '
            'bloqueante). tenant=%s ciclo=%s',
            len(advertencias), tenant.id, ciclo.id,
        )

    secciones = []
    for clave, nombre in _ATS_SECCIONES:
        filas = sorted(preguntas.get(clave, {}).values(), key=lambda f: f['orden'])
        n_seccion = len(aplicacion_ids_validas) if clave == 'D1' else len(aplicacion_ids_elegibles)
        secciones.append({'clave': clave, 'nombre': nombre, 'filas': filas, 'n': n_seccion})

    muestra = len(aplicacion_ids_validas)
    con_ats = len(aplicacion_ids_elegibles)
    clinica = ResultadoAplicacion.objects.filter(
        aplicacion__tenant=tenant, aplicacion__ciclo=ciclo,
        aplicacion__cuestionario__clave='I', requiere_atencion=True,
    ).count()

    return {
        'secciones': secciones,
        'advertencias_tecnicas': advertencias,
        'resumen': {
            'muestra':      muestra,
            'con_ats':      con_ats,
            'pct_ats':      round(con_ats / muestra * 100) if muestra else 0,
            'clinica':      clinica,
            'pct_clinica':  round(clinica / muestra * 100) if muestra else 0,
        },
    }


def _build_tasas_respuesta(tenant, ciclo):
    """Tasa de respuesta (completadas / aplicadas) por guía."""
    base = Aplicacion.objects.filter(tenant=tenant, ciclo=ciclo)
    etiquetas = [
        ('I',   'Guía I — Acontecimientos Traumáticos Severos'),
        ('III', 'Guía III — Factores de Riesgo Psicosocial'),
        ('V',   'Guía V — Datos del Trabajador'),
    ]
    tasas = []
    for clave, etiqueta in etiquetas:
        total = base.filter(cuestionario__clave=clave).count()
        if not total:
            continue
        comp = base.filter(cuestionario__clave=clave, estado='completado').count()
        tasas.append({
            'guia':        etiqueta,
            'total':       total,
            'completadas': comp,
            'pct':         round(comp / total * 100) if total else 0,
        })
    return tasas


URGENCIA_POR_NIVEL = {
    'muy_alto': 'Inmediata',
    'alto':     'A mediano plazo',
    'medio':    'Preventiva',
}


def _recomendacion_dominio(clave):
    """Extrae la acción concreta ('Se recomienda…') del texto clínico del dominio."""
    texto = interp.DOMINIO_INTERPRETACION.get(clave, {}).get('alto', '')
    idx = texto.find('Se recomienda')
    return texto[idx:] if idx != -1 else texto


def _recomendaciones_desde_dominios(dominios_prioritarios):
    """Genera recomendaciones concretas a partir de los dominios en nivel Alto / Muy alto,
    ordenadas por urgencia y citando el dominio que las origina."""
    orden = {'muy_alto': 0, 'alto': 1}
    recs = []
    for dom in sorted(dominios_prioritarios,
                      key=lambda d: (orden.get(d['categoria_predominante'], 9), -d['pct_promedio'])):
        nivel = dom['categoria_predominante']
        recs.append({
            'nivel':    nivel,
            'urgencia': URGENCIA_POR_NIVEL.get(nivel, 'A mediano plazo'),
            'texto':    _recomendacion_dominio(dom['clave']),
            'dominio':  f"{dom['clave']} — {dom['nombre']}",
        })
    return recs


def _recomendaciones_extendidas(dominios_oficiales_agregados):
    """Recomendaciones extensas (5-7 bullets por dominio, con intro), solo
    para el Informe Diagnóstico extendido (`informe_extendido=True`) — usa
    `norm.RECOMENDACIONES_DOMINIO` en vez de la extracción de una frase que
    consume el flujo borrador/aprobar. Se listan los dominios con intervención
    real (`pct_intervencion > 0`), de mayor a menor porcentaje."""
    recs = []
    for dom in sorted(dominios_oficiales_agregados, key=lambda d: -d['pct_intervencion']):
        if dom['pct_intervencion'] <= 0:
            continue
        contenido = norm.RECOMENDACIONES_DOMINIO.get(dom['nombre'])
        if not contenido:
            continue
        recs.append({
            'dominio':          dom['nombre'],
            'pct_intervencion': dom['pct_intervencion'],
            'intro':            contenido['intro'],
            'bullets':          contenido['bullets'],
        })
    return recs


# Variables demográficas (nombre en ReportDataNOM035) ↔ columna de las
# tablas de muestra del contexto, para aplicar la regla "variable con
# faltantes críticos → alerta, no gráfica".
_VARIABLE_DE_COLUMNA = {
    'Sexo':                                  'Sexo',
    'Grupo de edad':                         'Edad',
    'Estado civil':                          'Estado civil',
    'Nivel de estudios':                     'Nivel de estudios',
    'Tipo de puesto':                        'Tipo de puesto',
    'Tipo de contratación':                  'Tipo de contratación',
    'Tipo de personal':                      'Tipo de personal',
    'Tipo de jornada':                       'Tipo de jornada',
    'Realiza rotación':                      'Rotación de turnos',
    'Tiempo en el puesto actual':            'Tiempo en el puesto',
    'Tiempo de experiencia laboral total':   'Experiencia laboral',
}


def _dist_es_trivial(counts):
    """True si la distribución tiene una sola categoría con datos (o ninguna):
    una gráfica así no aporta información y se elimina del informe."""
    return sum(1 for c in counts if c) <= 1


def _build_graficas_informe(ctx):
    """PNGs (matplotlib) del Informe Diagnóstico extendido, derivados del
    mismo contexto que las tablas (fuente única ReportDataNOM035). Reglas de
    exclusión: no se grafican distribuciones de una sola categoría ni
    variables demográficas con faltantes críticos (esas llevan alerta)."""
    graficas = {}
    variables_sin_grafica = set(
        ctx.get('report_data', {}).get('exclusiones', {}).get('variables_sin_grafica', []))

    if ctx.get('ats_desglose'):
        graficas['ats'] = {
            seccion['clave']: graf.chart_ats(seccion['filas'])
            for seccion in ctx['ats_desglose']['secciones']
            if any(f['h_si'] or f['m_si'] for f in seccion['filas'])
        }

    dist_no_cero = [d for d in ctx['distribucion'] if d['count'] > 0]
    if len(dist_no_cero) > 1:
        graficas['distribucion'] = graf.chart_pie(
            [d['label'] for d in dist_no_cero],
            [d['count'] for d in dist_no_cero],
        )

    graficas['categorias'] = {
        c['nombre']: graf.chart_nivel_dist(c['nombre'], c['dist'], c['evaluados'])
        for c in ctx['categorias_agregadas']
        if not _dist_es_trivial(c['dist'].values())
    }
    graficas['dominios'] = {
        d['nombre']: graf.chart_nivel_dist(d['nombre'], d['dist'], d['evaluados'])
        for d in ctx['dominios_oficiales_agregados']
        if not _dist_es_trivial(d['dist'].values())
    }

    graficas['muestra'] = {}
    for tabla in ctx['muestra_tablas']:
        datos = tabla['datos']
        variable = _VARIABLE_DE_COLUMNA.get(tabla['col'], tabla['col'])
        if not datos or variable in variables_sin_grafica:
            continue  # faltantes críticos: alerta en el texto, sin gráfica
        counts = [d['count'] for d in datos]
        if _dist_es_trivial(counts):
            continue  # 100% en una sola respuesta: la gráfica no aporta
        labels = [d['label'] for d in datos]
        if len(datos) <= 3:
            graficas['muestra'][tabla['col']] = graf.chart_pie(labels, counts)
        else:
            total = sum(counts) or 1
            pcts = [c / total * 100 for c in counts]
            graficas['muestra'][tabla['col']] = graf.chart_barh(labels, pcts, counts)

    return graficas


def _riesgo_por_grupo(res_iii, keyfn):
    """Distribución de niveles individuales (Guía III) agrupada por un
    atributo del trabajador. NUNCA clasifica el promedio del grupo: reporta
    la moda, la distribución y el % en alto+muy alto. Los grupos con n menor
    al umbral de confidencialidad se reservan."""
    groups = {}
    for r in res_iii:
        k = keyfn(r.aplicacion.trabajador) or 'Sin dato'
        g = groups.setdefault(k, {'suma': 0, 'n': 0, 'dist': {x: 0 for x in DIST_KEYS}})
        g['suma'] += r.puntaje_total
        g['n'] += 1
        if r.categoria in g['dist']:
            g['dist'][r.categoria] += 1
    out = []
    for label, g in groups.items():
        if grupo_reservado(g['n']):
            out.append({
                'label':     label,
                'evaluados': g['n'],
                'reservado': True,
                'promedio':  None,
                'categoria': None,
                'categoria_label': ETIQUETA_RESERVADO,
                'pct_alto':  None,
                'dist':      None,
            })
            continue
        alto_mas = g['dist']['alto'] + g['dist']['muy_alto']
        cat_modal = max(DIST_KEYS, key=lambda k2: g['dist'][k2])
        out.append({
            'label':     label,
            'evaluados': g['n'],
            'reservado': False,
            'promedio':  round(g['suma'] / g['n']) if g['n'] else 0,  # descriptivo
            'categoria': cat_modal,  # moda de niveles individuales
            'categoria_label': CAT_LABELS[cat_modal],
            'pct_alto':  round(alto_mas / g['n'] * 100) if g['n'] else 0,
            'dist':      g['dist'],
        })
    out.sort(key=lambda x: -(x['pct_alto'] or 0))
    return out


# (_agregar_por_grupo fue sustituido por documents.report_data — fuente única.)


def _legacy_desde_report_data(filas_rd, orden_normativo):
    """Adapta las filas de ReportDataNOM035 (fuente única) al formato legado
    que consumen las gráficas, conclusiones y la plantilla HTML. NO recalcula
    nada: solo renombra campos."""
    idx = {nombre: i for i, nombre in enumerate(orden_normativo)}
    filas = sorted(filas_rd, key=lambda f: idx.get(f['nombre'], 999))
    return [
        {
            'nombre':                 f['nombre'],
            'evaluados':              f['n'],
            'dist':                   f['dist'],
            'pct_promedio':           round(f['pct_promedio'] or 0),
            'categoria_predominante': f['nivel_predominante'] or 'nulo',
            'requieren_atencion':     f['n_alto_o_muy_alto'],
            'pct_intervencion':       round(f['pct_alto_o_muy_alto'] or 0),
            'pct_accion':             round(f['pct_medio_o_mas'] or 0),
            'prioridad':              f['prioridad'],
        }
        for f in filas
    ]


def _agregar_dimensiones_oficiales(res_iii):
    """Agrega los registros `ResultadoDimension` (25 dimensiones, Tabla 6)
    del ciclo. SOLO estadísticos descriptivos: la NOM-035 no establece
    puntos de corte por dimensión, así que NUNCA se asigna nivel."""
    acc = {}
    for r in res_iii:
        for d in r.dimensiones.all():
            if not d.puntaje_max:
                continue
            e = acc.setdefault(d.orden, {
                'orden':           d.orden,
                'nombre':          d.nombre,
                'dominio_oficial': d.dominio_oficial,
                'pcts':            [],
                'pt':              0,
                'pm':              0,
            })
            e['pcts'].append(float(d.pct or 0))
            e['pt'] += d.puntaje
            e['pm'] += d.puntaje_max

    salida = []
    for e in sorted(acc.values(), key=lambda x: x['orden']):
        pcts = sorted(e['pcts'])
        n = len(pcts)
        mediana = pcts[n // 2] if n % 2 else (pcts[n // 2 - 1] + pcts[n // 2]) / 2
        salida.append({
            'orden':           e['orden'],
            'nombre':          e['nombre'],
            'dominio_oficial': e['dominio_oficial'],
            'n':               n,
            'pct_promedio':    round(sum(pcts) / n, 1),
            'pct_mediana':     round(mediana, 1),
            'pct_maximo':      round(pcts[-1], 1),
            'nota':            NOTA_DIMENSIONES,
        })
    return salida


# (_build_jerarquia_categorias fue sustituido por documents.report_data.)


def _build_conclusiones(ctx):
    """Texto y tablas de conclusiones (sección 10), generado a partir de los
    datos ya calculados del ciclo — mismo formato que el estándar LEAR
    (`generar_informe.py::build_seccion10`): calificación final + análisis
    por categoría y por dominio, rankeados de mayor a menor prioridad
    (% Alto+Muy alto, desempate por % Medio+Alto+Muy alto)."""
    dist = {d['key']: d['count'] for d in ctx['distribucion']}
    n = sum(dist.values())
    riesgo_alto   = dist.get('alto', 0) + dist.get('muy_alto', 0)
    accion_global = dist.get('medio', 0) + riesgo_alto
    pct_riesgo  = round(riesgo_alto / n * 100, 1) if n else 0
    pct_accion  = round(accion_global / n * 100, 1) if n else 0
    nivel_predominante = max(DIST_KEYS, key=lambda k: dist.get(k, 0)) if n else 'nulo'

    categorias_rankeadas = sorted(
        ctx['categorias_agregadas'],
        key=lambda x: (x['pct_intervencion'], x['pct_accion']), reverse=True)
    dominios_rankeados = sorted(
        ctx['dominios_oficiales_agregados'],
        key=lambda x: (x['pct_intervencion'], x['pct_accion']), reverse=True)

    return {
        'muestra':               n,
        'poblacion':             ctx['tenant'].num_trabajadores,
        'riesgo_alto':           riesgo_alto,
        'pct_riesgo':            pct_riesgo,
        'accion_global':         accion_global,
        'pct_accion':            pct_accion,
        'nivel_predominante':    nivel_predominante,
        'categorias_rankeadas':  categorias_rankeadas,
        'dominios_rankeados':    dominios_rankeados,
        # Top 4 — se conservan para el borrador (bullets, sin cambios).
        'categorias_destacadas': [c for c in categorias_rankeadas if c['pct_intervencion'] > 0][:4],
        'dominios_destacados':   [d for d in dominios_rankeados if d['pct_intervencion'] > 0][:4],
    }


def _build_anexos(res_iii, res_i):
    """Tablas de anexos (13.1-13.4), siempre identificadas por el número de
    empleado del trabajador (Guía V) — independientemente del modo anónimo
    del cuerpo del reporte, así se presentan en el estándar de referencia."""
    claves = _clave_map(list(res_iii) + list(res_i))
    valoracion_clinica = {r.aplicacion.trabajador_id: r.requiere_atencion for r in res_i}
    ats_reportado = {
        rd.resultado.aplicacion.trabajador_id: rd.puntaje > 0
        for rd in ResultadoDominio.objects.filter(resultado__in=res_i, dominio__clave='D1')
            .select_related('resultado__aplicacion')
    }

    ats_final = []
    for r in sorted(res_iii, key=lambda x: -x.puntaje_total):
        tid = r.aplicacion.trabajador_id
        ats_final.append({
            'clave':              claves.get(tid, '—'),
            'ats':                ('Sí' if ats_reportado.get(tid)
                                    else 'No' if tid in ats_reportado else 'No aplicado'),
            'valoracion_clinica': ('Requerida' if valoracion_clinica.get(tid)
                                    else 'Sin indicadores' if tid in valoracion_clinica else 'No aplicado'),
            # Columna combinada — se conserva solo para el borrador, que no
            # separa ATS reportado de valoración clínica.
            'ats_combinado':      ('Atención requerida' if valoracion_clinica.get(tid)
                                    else 'Sin indicadores' if tid in valoracion_clinica else 'No aplicado'),
            'calificacion_final': r.puntaje_total,
            'nivel_final':        r.categoria,
        })

    orden_categorias = [c for c, _ in _CATEGORIA_DOMINIOS]
    orden_dominios = [dom for _, doms in _CATEGORIA_DOMINIOS for dom in doms]

    categorias_filas, dominios_filas = [], []
    for r in sorted(res_iii, key=lambda x: claves.get(x.aplicacion.trabajador_id, '')):
        tid = r.aplicacion.trabajador_id
        cat_acc = defaultdict(lambda: {'pt': 0, 'pm': 0})
        dom_acc = defaultdict(lambda: {'pt': 0, 'pm': 0})
        for d in r.dominios_oficiales.all():
            if not d.puntaje_max:
                continue
            categoria = _CATEGORIA_DE_DOMINIO.get(d.nombre)
            if categoria:
                cat_acc[categoria]['pt'] += d.puntaje
                cat_acc[categoria]['pm'] += d.puntaje_max
            dom_acc[d.nombre]['pt'] += d.puntaje
            dom_acc[d.nombre]['pm'] += d.puntaje_max

        def _nivel(acc, nombre, cortes):
            v = acc.get(nombre)
            if not v or not v['pm']:
                return None
            return _categoria_por_rangos(v['pt'], cortes)

        categorias_filas.append({
            'clave':   claves.get(tid, '—'),
            'niveles': [_nivel(cat_acc, c, _CORTES_CATEGORIA[c]) for c in orden_categorias],
        })
        dominios_filas.append({
            'clave':   claves.get(tid, '—'),
            'niveles': [_nivel(dom_acc, d, _CORTES_DOMINIO[d]) for d in orden_dominios],
        })

    return {
        'ats_final':           ats_final,
        'categorias_columnas': orden_categorias,
        'categorias_filas':    categorias_filas,
        'dominios_columnas':   orden_dominios,
        'dominios_filas':      dominios_filas,
        'agrupacion':          norm.TABLA_AGRUPACION_DOMINIOS,
    }


def _build_psico_context(tenant, ciclo, resultados_qs, anonimo, informe_extendido=False,
                         incluir_anexo_confidencial=True):
    """`informe_extendido=True` agrega gráficas (matplotlib) y recomendaciones
    extensas por dominio — solo lo activa `descargar_informe_diagnostico`
    (botón "Descargar informe"). El flujo borrador/aprobar existente nunca lo
    pasa, así que su contenido/comportamiento queda intacto."""
    ctx = _build_context(tenant, ciclo, resultados_qs, anonimo=anonimo)

    resultados_lista = list(resultados_qs.select_related(
        'aplicacion__trabajador', 'aplicacion__cuestionario'
    ).prefetch_related('dominios__dominio', 'dominios_oficiales', 'categorias', 'dimensiones'))
    res_iii = [r for r in resultados_lista
               if r.aplicacion.cuestionario.clave == 'III'
               and getattr(r, 'estatus_validacion', 'valido') == 'valido']
    res_i = [r for r in resultados_lista
             if r.aplicacion.cuestionario.clave == 'I'
             and getattr(r, 'estatus_validacion', 'valido') == 'valido']
    riesgo_por_grupo = {
        'sexo':    _riesgo_por_grupo(res_iii, lambda t: t.get_sexo_display() if t.sexo else None),
        'edad':    _riesgo_por_grupo(res_iii, lambda t: _grupo_edad(t.edad)),
        'puesto':  _riesgo_por_grupo(res_iii, lambda t: t.tipo_puesto or None),
        'jornada': _riesgo_por_grupo(res_iii, lambda t: t.get_tipo_jornada_display() if t.tipo_jornada else None),
    }

    # Interpretación clínica del nivel de riesgo global (versión extendida)
    ctx['nivel_global_texto'] = interp.NIVEL_GLOBAL[ctx['nivel_global']]

    # Enriquecer cada dominio con qué mide + interpretación clínica si es prioritario
    for dom in ctx['dominios_agregados']:
        meta = interp.DOMINIO_INTERPRETACION.get(dom['clave'], {})
        dom['mide'] = meta.get('mide', '')
        prioritario = dom['categoria_predominante'] in ('alto', 'muy_alto')
        dom['prioritario'] = prioritario
        dom['interpretacion'] = meta.get('alto', '') if prioritario else ''

    dominios_prioritarios = [d for d in ctx['dominios_agregados'] if d['prioritario']]

    # Bloques internos de captura (D1-D14): análisis complementario NO
    # normativo — la NOM no define cortes por bloque. Se conservan para el
    # detalle técnico, claramente etiquetados.
    bloques_captura = [
        {**dom, 'dominio_oficial': norm.DOMINIO_OFICIAL_DE_BLOQUE.get(dom['clave'], dom['nombre'])}
        for dom in ctx['dominios_agregados']
    ]

    # Dimensiones OFICIALES (25, Tabla 6): indicador exclusivamente
    # descriptivo — la NOM-035 no establece puntos de corte por dimensión.
    dimensiones = _agregar_dimensiones_oficiales(res_iii)

    # ------------------------------------------------------------------
    # MOTOR DE COMPOSICIÓN EJECUTIVA: objeto único ReportDataNOM035.
    # Todas las tarjetas, tablas y rankings del informe derivan de aquí;
    # los agregados "legado" que consumen gráficas/plantillas se adaptan
    # con _legacy_desde_report_data (sin recalcular).
    # ------------------------------------------------------------------
    report_data = build_report_data(tenant, ciclo)

    orden_categorias = [c for c, _ in _CATEGORIA_DOMINIOS]
    orden_dominios = [dom for _, doms in _CATEGORIA_DOMINIOS for dom in doms]
    categorias_agregadas = _legacy_desde_report_data(
        report_data['tablas']['categorias']['filas'], orden_categorias)
    dominios_oficiales_agregados = _legacy_desde_report_data(
        report_data['tablas']['dominios']['filas'], orden_dominios)

    # Guía I — texto interpretativo según haya o no casos positivos
    if ctx['guia_i']['requieren_atencion'] > 0:
        guia_i_texto = interp.GUIA_I_TEXTO_POSITIVO
    else:
        guia_i_texto = interp.GUIA_I_TEXTO_SIN_INDICADORES

    def _tabla_muestra(titulo, col, dist):
        return {
            'titulo': titulo, 'col': col,
            'datos': dist['datos'],
            'n_total': dist['n_total'], 'n_valido': dist['n_valido'],
            'n_faltante': dist['n_faltante'], 'pct_faltante': dist['pct_faltante'],
        }

    muestra = _build_muestra(tenant, ciclo)
    muestra_tablas = [
        _tabla_muestra('Distribución por sexo',            'Sexo',             muestra['sexo']),
        _tabla_muestra('Distribución por grupo de edad',   'Grupo de edad',    muestra['edad']),
        _tabla_muestra('Distribución por estado civil',    'Estado civil',     muestra['estado_civil']),
        _tabla_muestra('Distribución por nivel de estudios','Nivel de estudios',muestra['estudios']),
        _tabla_muestra('Distribución por tipo de puesto',  'Tipo de puesto',   muestra['puesto']),
        _tabla_muestra('Distribución por tipo de contratación', 'Tipo de contratación', muestra['contratacion']),
        _tabla_muestra('Distribución por tipo de personal','Tipo de personal', muestra['personal']),
        _tabla_muestra('Distribución por tipo de jornada', 'Tipo de jornada',  muestra['jornada']),
        _tabla_muestra('Rotación de turnos',               'Realiza rotación', muestra['rotacion']),
        _tabla_muestra('Tiempo en el puesto actual',       'Tiempo en el puesto actual', muestra['tiempo_puesto']),
        _tabla_muestra('Tiempo de experiencia laboral total', 'Tiempo de experiencia laboral total', muestra['experiencia_total']),
    ]
    riesgo_tablas = [
        {'titulo': 'Por sexo',           'col': 'Sexo',          'datos': riesgo_por_grupo['sexo']},
        {'titulo': 'Por grupo de edad',  'col': 'Grupo de edad', 'datos': riesgo_por_grupo['edad']},
        {'titulo': 'Por tipo de puesto', 'col': 'Tipo de puesto','datos': riesgo_por_grupo['puesto']},
        {'titulo': 'Por tipo de jornada','col': 'Tipo de jornada','datos': riesgo_por_grupo['jornada']},
    ]

    sexo_dist = {d['label']: d['count'] for d in muestra['sexo']['datos']}

    ctx.update({
        'report_data':           report_data,
        'anonimo':               anonimo,
        # Se llena solo al aprobar, vía _inyectar_validacion_revisor() — nunca es texto libre.
        'validacion':            None,
        'marco':                 interp.MARCO_NORMATIVO,
        'guia_i_definicion':     interp.GUIA_I_DEFINICION,
        'guia_i_texto':          guia_i_texto,
        'limitaciones':          interp.LIMITACIONES,
        'muestra':               muestra,
        'muestra_tablas':        muestra_tablas,
        'tasas_respuesta':       _build_tasas_respuesta(tenant, ciclo),
        'dominios_prioritarios': dominios_prioritarios,
        'dimensiones':           dimensiones,       # 25 oficiales, descriptivas
        'bloques_captura':       bloques_captura,   # D1-D14, no normativos
        'nota_dimensiones':      NOTA_DIMENSIONES,
        'categorias_agregadas':  categorias_agregadas,
        'dominios_oficiales_agregados': dominios_oficiales_agregados,
        'riesgo_por_grupo':      riesgo_por_grupo,
        'riesgo_tablas':         riesgo_tablas,
        # Recomendaciones derivadas de los dominios prioritarios (sustituye las genéricas)
        'recomendaciones':       _recomendaciones_desde_dominios(dominios_prioritarios),
        # ---- Datos del centro de trabajo: domicilio y giro SIEMPRE del
        # tenant (hallazgo H-17 de la auditoría — un informe jamás debe salir
        # con el domicilio o el giro de otra planta). La razón social es fija
        # porque todas las plantas son la misma persona moral. ----
        'datos_centro_trabajo': {
            'nombre':    RAZON_SOCIAL_CORPORATIVA,
            'direccion': tenant.direccion,
            'giro':      tenant.giro or '—',
        },
        'objetivo_general':      norm.OBJETIVO_GENERAL,
        'objetivos_especificos': norm.OBJETIVOS_ESPECIFICOS,
        'definiciones':          norm.DEFINICIONES,
        'justificacion_muestra': {
            'intro':     norm.JUSTIFICACION_MUESTRA_INTRO,
            'poblacion': tenant.num_trabajadores,
            'muestra':   ctx['resumen']['total_guia_iii'],
            'hombres':   sexo_dist.get('Masculino', 0),
            'mujeres':   sexo_dist.get('Femenino', 0),
        },
        'metodologia': norm.METODOLOGIA,
    })
    ctx['conclusiones'] = _build_conclusiones(ctx)
    # Anexo con resultados individuales (folios): es información CONFIDENCIAL
    # y no debe integrarse automáticamente a la versión general del informe.
    ctx['anexos'] = _build_anexos(res_iii, res_i) if incluir_anexo_confidencial else None
    ctx['incluye_anexo_confidencial'] = bool(incluir_anexo_confidencial)

    if informe_extendido:
        ctx['acciones_generales'] = norm.ACCIONES_GENERALES
        ctx['recomendaciones_extendidas'] = _recomendaciones_extendidas(dominios_oficiales_agregados)
        ctx['justificacion_muestra']['parrafos'] = norm.JUSTIFICACION_MUESTRA_PARRAFOS
        ctx['justificacion_muestra']['ecuacion1'] = _datos_muestra_ecuacion1(tenant, ciclo)
        # Fuente única: el flujo de muestra viene de ReportDataNOM035.
        ctx['flujo_muestra'] = report_data['tablas']['flujo_muestra']['filas']
        ctx['ats_desglose'] = _ats_desglose(tenant, ciclo)
        ctx['graficas'] = _build_graficas_informe(ctx)

    return ctx


def _inyectar_validacion_revisor(ctx, revisor):
    """Sobreescribe la sección de cierre con los datos reales del revisor que
    aprueba el reporte. Solo se llama al aprobar — el borrador nunca lleva
    estos datos."""
    ctx['validacion'] = {
        'nombre_revisor':   revisor.nombre_completo,
        'cedula_revisor':   revisor.cedula_profesional,
        'fecha_aprobacion': _fecha_es(timezone.localdate()),
    }
    return ctx


def _resultados_qs_o_none(tenant, ciclo):
    qs = ResultadoAplicacion.objects.filter(aplicacion__tenant=tenant, aplicacion__ciclo=ciclo)
    return qs if qs.exists() else None


@api_view(['POST'])
@permission_classes([IsSuperAdmin])
def reporte_psicologico_borrador(request):
    ciclo_id = request.data.get('ciclo_id') or request.query_params.get('ciclo_id')
    if not ciclo_id:
        return _wrap(None, errors={'detail': 'ciclo_id requerido'}, status_code=400)

    anonimo_raw = request.data.get('anonimo', request.query_params.get('anonimo', False))
    anonimo = str(anonimo_raw).lower() in ('true', '1', 'yes')

    try:
        ciclo = CicloNOM.objects.select_related('tenant').get(id=ciclo_id)
    except CicloNOM.DoesNotExist:
        return _wrap(None, errors={'detail': 'Ciclo no encontrado'}, status_code=404)
    tenant = ciclo.tenant

    reporte, _created = ReportePsicologico.objects.get_or_create(
        tenant=tenant, ciclo=ciclo, defaults={'estado': ReportePsicologico.ESTADO_BORRADOR},
    )
    if reporte.is_aprobado:
        return _wrap(None, errors={'detail': 'El reporte ya fue aprobado y es inmutable.'}, status_code=409)

    resultados_qs = _resultados_qs_o_none(tenant, ciclo)
    if resultados_qs is None:
        return _wrap(None, errors={'detail': 'No hay resultados calculados para este ciclo.'}, status_code=400)

    ctx = _build_psico_context(tenant, ciclo, resultados_qs, anonimo)

    reporte.estado = ReportePsicologico.ESTADO_BORRADOR
    reporte.anonimo = anonimo
    reporte.generado_por = request.user
    reporte.generado_en = timezone.now()
    reporte.save()

    docx_buf = build_reporte_psicologico_docx(ctx)
    filename = f'borrador_reporte_psicologico_{tenant.nombre.replace(" ", "_")}_{ciclo.anio}.docx'
    response = HttpResponse(
        docx_buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['POST'])
@permission_classes([IsSuperAdmin])
def reporte_psicologico_aprobar(request):
    ciclo_id = request.data.get('ciclo_id') or request.query_params.get('ciclo_id')
    if not ciclo_id:
        return _wrap(None, errors={'detail': 'ciclo_id requerido'}, status_code=400)

    try:
        ciclo = CicloNOM.objects.select_related('tenant').get(id=ciclo_id)
    except CicloNOM.DoesNotExist:
        return _wrap(None, errors={'detail': 'Ciclo no encontrado'}, status_code=404)
    tenant = ciclo.tenant

    if tenant.consultor_id != request.user.id:
        return _wrap(
            None,
            errors={'detail': 'Solo el consultor asignado a esta empresa puede aprobar su reporte.'},
            status_code=403,
        )
    if not request.user.cedula_profesional.strip():
        return _wrap(
            None,
            errors={'detail': 'No tienes una cédula profesional registrada. Contacta al administrador del sistema.'},
            status_code=403,
        )

    try:
        reporte = ReportePsicologico.objects.get(tenant=tenant, ciclo=ciclo)
    except ReportePsicologico.DoesNotExist:
        return _wrap(
            None,
            errors={'detail': 'No existe un borrador para este ciclo. Genera el borrador primero.'},
            status_code=404,
        )
    if reporte.is_aprobado:
        return _wrap(None, errors={'detail': 'El reporte ya fue aprobado previamente.'}, status_code=409)

    resultados_qs = _resultados_qs_o_none(tenant, ciclo)
    if resultados_qs is None:
        return _wrap(None, errors={'detail': 'No hay resultados calculados para este ciclo.'}, status_code=400)

    ctx = _build_psico_context(tenant, ciclo, resultados_qs, reporte.anonimo)

    # Bloqueo de emisión definitiva: no se aprueba un reporte con errores
    # críticos de consistencia detectados por el motor de composición.
    validaciones = ctx['report_data']['validaciones']
    if not validaciones['puede_emitirse']:
        return _wrap(
            None,
            errors={'detail': 'No es posible aprobar el reporte. Errores críticos: '
                              + ' | '.join(validaciones['errores_criticos'])},
            status_code=409,
        )

    ctx = _inyectar_validacion_revisor(ctx, request.user)
    html_final = render_to_string('documents/reporte_psicologico.html', ctx, request=request)

    reporte.estado = ReportePsicologico.ESTADO_APROBADO
    reporte.revisor = request.user
    reporte.nombre_revisor = request.user.nombre_completo
    reporte.cedula_revisor = request.user.cedula_profesional
    reporte.fecha_aprobacion = timezone.now()
    reporte.contenido_html_final = html_final
    reporte.save()

    return _wrap({
        'estado':           reporte.estado,
        'nombre_revisor':   reporte.nombre_revisor,
        'fecha_aprobacion': reporte.fecha_aprobacion,
    })


@api_view(['GET'])
@permission_classes([IsTenantAdmin])
def reporte_psicologico_descargar(request):
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return HttpResponse('ciclo_id requerido', status=400)

    tenant = _tenant_para_ciclo(request, ciclo_id)
    try:
        ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
    except CicloNOM.DoesNotExist:
        return HttpResponse('Ciclo no encontrado', status=404)

    try:
        reporte = ReportePsicologico.objects.get(tenant=tenant, ciclo=ciclo)
    except ReportePsicologico.DoesNotExist:
        return HttpResponse('Aún no se ha generado un reporte psicológico para este ciclo.', status=404)

    if not reporte.is_aprobado:
        return HttpResponse('El reporte psicológico está en revisión y aún no ha sido aprobado.', status=409)

    filename = f'reporte_psicologico_{tenant.nombre.replace(" ", "_")}_{ciclo.anio}.pdf'
    return _pdf_response(reporte.contenido_html_final, filename, request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_psicologico_estado(request):
    ciclo_id = request.query_params.get('ciclo_id')
    if not ciclo_id:
        return _wrap(None, errors={'detail': 'ciclo_id requerido'}, status_code=400)

    if request.user.is_super_admin:
        try:
            ciclo = CicloNOM.objects.select_related('tenant').get(id=ciclo_id)
        except CicloNOM.DoesNotExist:
            return _wrap(None, errors={'detail': 'Ciclo no encontrado'}, status_code=404)
        tenant = ciclo.tenant
    else:
        tenant = request.user.tenant
        try:
            ciclo = CicloNOM.objects.get(id=ciclo_id, tenant=tenant)
        except CicloNOM.DoesNotExist:
            return _wrap(None, errors={'detail': 'Ciclo no encontrado'}, status_code=404)

    reporte = ReportePsicologico.objects.filter(tenant=tenant, ciclo=ciclo).first()
    puede_aprobar = bool(
        request.user.is_super_admin
        and tenant.consultor_id == request.user.id
        and request.user.cedula_profesional.strip()
    )
    return _wrap({
        'existe':                 reporte is not None,
        'estado':                 reporte.estado if reporte else None,
        'anonimo':                reporte.anonimo if reporte else None,
        'generado_en':            reporte.generado_en if reporte else None,
        'nombre_revisor':         reporte.nombre_revisor if reporte else '',
        'fecha_aprobacion':       reporte.fecha_aprobacion if reporte else None,
        'puede_generar_borrador': bool(request.user.is_super_admin),
        'puede_aprobar':          puede_aprobar,
    })
