"""
Reimporta las respuestas de Guía V corregidas por el cliente.

Cuando una parte de las respuestas de Guía V llega en un formato que el
informe no puede interpretar ("si", "tres años", "desde 2023"), se exporta
un Excel con esas filas marcadas, el cliente vuelve a preguntar y devuelve
el archivo. Este comando toma el archivo devuelto y actualiza las
`RespuestaPregunta` correspondientes.

Reglas:
  * Solo escribe donde el valor del archivo difiere del que está en la base.
  * Una celda vacía significa "sin cambio", nunca borra una respuesta.
  * Valida antes de escribir: las preguntas de opción contra el catálogo de
    la pregunta, y las dos de duración contra `_norm_experiencia`, que es el
    mismo normalizador que decide si el informe cuenta el dato o lo reporta
    como faltante.
  * Sin --apply no toca la base: imprime el diff y termina.

Uso:
    python manage.py importar_guia_v_corregida archivo.xlsx --ciclo 7
    python manage.py importar_guia_v_corregida archivo.xlsx --ciclo 7 --apply
"""

import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from m00_onboarding.models import CicloNOM, Trabajador
from m00_onboarding.importador import _norm_edad, _norm_experiencia
from m05_questionnaires.models import Aplicacion, Pregunta, RespuestaPregunta

HOJA_POR_DEFECTO = 'Guia V - Respuestas'

# Encabezados esperados de la hoja, en orden. Los primeros cuatro son
# identificación; del quinto en adelante son las 15 preguntas de Guía V,
# en el mismo orden que `Pregunta.orden`.
COLS_ID = ['id interno', 'no. empleado', 'nombre (perfil rh)', 'estado guia v']
COLS_PREGUNTA = [
    'nombre completo',
    'fecha (dia / mes / ano)',
    'sexo',
    'edad (anos completos)',
    'estado civil',
    'ultimo nivel de estudios',
    'ocupacion / profesion / puesto',
    'departamento / seccion / area',
    'tipo de puesto',
    'tipo de contratacion',
    'tipo de personal',
    'tipo de jornada',
    'rotacion de turnos',
    'experiencia en el puesto actual',
    'tiempo de experiencia laboral total',
]

ORDEN_ROTACION = 13
ORDEN_EDAD = 4
ORDENES_DURACION = (14, 15)

SI_NO = {'si': 1, 'sí': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0}


def _norm(texto):
    """minúsculas, sin acentos y con espacios colapsados."""
    s = unicodedata.normalize('NFKD', str(texto or ''))
    s = s.encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().lower()


class Command(BaseCommand):
    help = 'Reimporta respuestas de Guía V corregidas desde el Excel devuelto por el cliente.'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta del .xlsx devuelto por el cliente')
        parser.add_argument('--ciclo', type=int, required=True,
                            help='ID del CicloNOM al que pertenecen las aplicaciones')
        parser.add_argument('--hoja', default=HOJA_POR_DEFECTO,
                            help=f'Nombre de la hoja a leer (por defecto "{HOJA_POR_DEFECTO}")')
        parser.add_argument('--apply', action='store_true',
                            help='Escribe los cambios. Sin esta bandera solo muestra el diff.')

    # -- lectura ----------------------------------------------------------

    def _abrir_hoja(self, archivo, hoja):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Falta openpyxl (pip install openpyxl)')
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'No existe el archivo: {archivo}')
        if hoja not in wb.sheetnames:
            raise CommandError(f'El archivo no tiene la hoja "{hoja}". Tiene: {wb.sheetnames}')
        ws = wb[hoja]

        encabezados = [_norm(c.value) for c in ws[1]]
        esperados = COLS_ID + COLS_PREGUNTA
        if encabezados[:len(esperados)] != esperados:
            faltan = [e for e in esperados if e not in encabezados]
            raise CommandError(
                'Los encabezados de la hoja no son los del archivo exportado. '
                f'No coinciden / faltan: {faltan or encabezados[:len(esperados)]}'
            )
        return ws

    # -- validación por pregunta -----------------------------------------

    def _valor_para(self, pregunta, celda):
        """(valor, valor_texto, error) para el contenido de una celda.

        Devuelve error != None cuando el dato no se puede guardar en el
        formato que el informe sabe leer.
        """
        crudo = '' if celda is None else str(celda).strip()
        if not crudo:
            return None, None, None  # celda vacía → sin cambio

        if pregunta.orden == ORDEN_ROTACION:
            v = SI_NO.get(_norm(crudo))
            if v is None:
                return None, None, f'se esperaba Si o No, llegó {crudo!r}'
            return v, '', None

        if pregunta.orden == ORDEN_EDAD:
            if _norm_edad(crudo) is None:
                return None, None, f'no se puede leer como edad: {crudo!r}'
            return None, crudo, None

        if pregunta.orden in ORDENES_DURACION:
            if _norm_experiencia(crudo) is None:
                return None, None, f'no se puede leer como duración: {crudo!r}'
            return None, crudo, None

        if pregunta.tipo_respuesta == 'opcion' and pregunta.opciones:
            catalogo = {_norm(o): o for o in pregunta.opciones}
            opcion = catalogo.get(_norm(crudo))
            if opcion is None:
                return None, None, (f'{crudo!r} no está en las opciones de la pregunta '
                                    f'({", ".join(pregunta.opciones)})')
            return None, opcion, None

        return None, crudo, None

    # -- ejecución --------------------------------------------------------

    def handle(self, *args, **opts):
        ws = self._abrir_hoja(opts['archivo'], opts['hoja'])

        try:
            ciclo = CicloNOM.objects.select_related('tenant').get(id=opts['ciclo'])
        except CicloNOM.DoesNotExist:
            raise CommandError(f'No existe el ciclo {opts["ciclo"]}')
        tenant = ciclo.tenant

        preguntas = {
            p.orden: p
            for p in Pregunta.objects.filter(dominio__cuestionario__clave='V')
                                     .select_related('dominio')
        }
        if len(preguntas) != len(COLS_PREGUNTA):
            raise CommandError(
                f'La Guía V tiene {len(preguntas)} preguntas en la base y el archivo '
                f'trae {len(COLS_PREGUNTA)} columnas de pregunta.'
            )

        aplicaciones = {
            a.trabajador_id: a
            for a in Aplicacion.objects.filter(
                tenant=tenant, ciclo=ciclo, cuestionario__clave='V',
            ).select_related('trabajador')
        }
        respuestas = {}
        for r in RespuestaPregunta.objects.filter(aplicacion__in=aplicaciones.values()):
            respuestas[(r.aplicacion_id, r.pregunta_id)] = r

        cambios, nuevas, errores, sin_match = [], [], [], []

        for fila in range(2, ws.max_row + 1):
            id_interno = ws.cell(fila, 1).value
            num_empleado = ws.cell(fila, 2).value
            try:
                id_interno = int(str(id_interno).strip())
            except (TypeError, ValueError):
                id_interno = None
            num_empleado = str(num_empleado).strip() if num_empleado is not None else ''
            if id_interno is None and not num_empleado:
                # fila vacía o la leyenda al pie de la hoja
                continue

            trabajador = None
            if id_interno is not None:
                trabajador = Trabajador.objects.filter(tenant=tenant, id=id_interno).first()
            if trabajador is None and num_empleado:
                trabajador = Trabajador.objects.filter(
                    tenant=tenant, num_empleado=num_empleado,
                ).first()
            if trabajador is None:
                sin_match.append((fila, id_interno, num_empleado, 'no existe el trabajador'))
                continue
            if num_empleado and trabajador.num_empleado and \
                    num_empleado != trabajador.num_empleado:
                sin_match.append((fila, id_interno, num_empleado,
                                  f'el ID interno corresponde a otro empleado '
                                  f'({trabajador.num_empleado})'))
                continue

            aplicacion = aplicaciones.get(trabajador.id)
            if aplicacion is None:
                sin_match.append((fila, id_interno, num_empleado,
                                  'sin aplicación de Guía V en este ciclo'))
                continue

            for i, _ in enumerate(COLS_PREGUNTA):
                orden = i + 1
                pregunta = preguntas[orden]
                celda = ws.cell(fila, len(COLS_ID) + 1 + i).value
                valor, valor_texto, error = self._valor_para(pregunta, celda)
                if error:
                    errores.append((fila, trabajador, orden, error))
                    continue
                if valor is None and valor_texto is None:
                    continue

                actual = respuestas.get((aplicacion.id, pregunta.id))
                if actual is None:
                    nuevas.append((trabajador, aplicacion, pregunta, valor, valor_texto))
                elif (actual.valor, actual.valor_texto or '') != (valor, valor_texto or ''):
                    cambios.append((trabajador, actual, pregunta, valor, valor_texto))

        self._reportar(cambios, nuevas, errores, sin_match)

        if not opts['apply']:
            self.stdout.write(self.style.WARNING(
                '\nSimulación: no se escribió nada. Repite con --apply para aplicar.'))
            return
        if not cambios and not nuevas:
            self.stdout.write('\nNada que aplicar.')
            return

        self._aplicar(cambios, nuevas, preguntas)

    # -- salida -----------------------------------------------------------

    def _reportar(self, cambios, nuevas, errores, sin_match):
        if sin_match:
            self.stdout.write(self.style.ERROR(f'\nFilas sin trabajador ({len(sin_match)}):'))
            for fila, id_i, emp, motivo in sin_match:
                self.stdout.write(f'  fila {fila}: id={id_i} emp={emp} — {motivo}')

        if errores:
            self.stdout.write(self.style.WARNING(
                f'\nRespuestas que siguen sin ser interpretables ({len(errores)}) '
                '— se omiten, el resto de la fila sí se procesa:'))
            for fila, trabajador, orden, error in errores:
                self.stdout.write(f'  fila {fila} [{trabajador.num_empleado}] P{orden}: {error}')

        if nuevas:
            self.stdout.write(self.style.SUCCESS(f'\nRespuestas nuevas ({len(nuevas)}):'))
            for trabajador, _, pregunta, valor, valor_texto in nuevas:
                v = valor_texto if valor is None else ('Si' if valor else 'No')
                self.stdout.write(f'  [{trabajador.num_empleado}] P{pregunta.orden} = {v!r}')

        if cambios:
            self.stdout.write(self.style.SUCCESS(f'\nRespuestas modificadas ({len(cambios)}):'))
            for trabajador, actual, pregunta, valor, valor_texto in cambios:
                antes = actual.valor_texto if actual.valor is None else (
                    'Si' if actual.valor else 'No')
                despues = valor_texto if valor is None else ('Si' if valor else 'No')
                self.stdout.write(
                    f'  [{trabajador.num_empleado}] P{pregunta.orden}: {antes!r} -> {despues!r}')

        self.stdout.write(
            f'\nResumen: {len(nuevas)} nuevas, {len(cambios)} modificadas, '
            f'{len(errores)} sin interpretar, {len(sin_match)} filas sin trabajador.')

    # -- escritura --------------------------------------------------------

    @transaction.atomic
    def _aplicar(self, cambios, nuevas, preguntas):
        for _, actual, _, valor, valor_texto in cambios:
            actual.valor = valor
            actual.valor_texto = valor_texto or ''
            actual.save(update_fields=['valor', 'valor_texto'])

        afectadas = {a for _, a, _, _, _ in nuevas}
        for _, aplicacion, pregunta, valor, valor_texto in nuevas:
            RespuestaPregunta.objects.create(
                aplicacion=aplicacion, pregunta=pregunta, tenant=aplicacion.tenant,
                valor=valor, valor_texto=valor_texto or '',
            )

        completadas = 0
        for aplicacion in afectadas:
            if aplicacion.estado == 'completado':
                continue
            if aplicacion.respuestas.count() >= len(preguntas):
                aplicacion.marcar_completado()
                completadas += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nAplicado: {len(cambios)} modificadas, {len(nuevas)} nuevas, '
            f'{completadas} aplicaciones marcadas como completadas.'))
